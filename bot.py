import logging
import os
import re
from io import BytesIO
from datetime import datetime
from unidecode import unidecode
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ConversationHandler, filters, ContextTypes
)
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from googletrans import Translator
import math
import asyncio




    # =================== Загрузка переменных ===================
load_dotenv()
TOKEN = os.getenv("TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

# =================== Логирование ===================
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# =================== Состояния ===================
SERIAL, ALLOCATION, TEAM_NUMBER, USER, DESCRIPTION = range(5)
translator = Translator()



# =================== Вспомогательные функции ===================
def get_workbook(report_type="LDR"):
    current_dir = os.path.dirname(__file__)
    if report_type.upper() == "MFR":
        filename = "MFR.xlsx"
    else:
        filename = "LDR.xlsx"
    return load_workbook(os.path.join(current_dir, "excel", filename))


def get_logo_bytes():
    current_dir = os.path.dirname(__file__)
    with open(os.path.join(current_dir, "logo", "Drive the NPA way.png"), "rb") as f:
        return BytesIO(f.read())

async def translate_to_en(text: str) -> str:
    result = await translator.translate(text, dest='en')
    return result.text




def set_cell(ws, cell, value):
    try:
        ws[cell].value = value
    except AttributeError:
        # если попали в объединённую ячейку, ищем верхнюю левую
        for merged_range in ws.merged_cells.ranges:
            if cell in merged_range:
                top_left = merged_range.min_row, merged_range.min_col
                ws.cell(row=top_left[0], column=top_left[1], value=value)
                break
def auto_adjust(ws, cells):
    for cell in cells:
        value = ws[cell].value
        if value:
            col_letter = ''.join(filter(str.isalpha, cell))
            ws.column_dimensions[col_letter].width = max(
                ws.column_dimensions[col_letter].width or 10,
                len(str(value)) + 2
            )
            ws.row_dimensions[ws[cell].row].height = max(
                ws.row_dimensions[ws[cell].row].height or 15,
                15
            )



# =================== Главное меню ===================
async def main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("LDR (Lost / Damage) | Втрачено або пошкоджено", callback_data="ldr")],
        [InlineKeyboardButton("MFR (Mechanical failure) | Механічне пошкодження авто", callback_data="mfr")],
        [InlineKeyboardButton("Monthly Inspection Form | Щомісячний огляд авто", callback_data="monthly_form")],
        [InlineKeyboardButton("🚨 Порядок дій при ДТП | Accident Procedure", callback_data="accident_procedure")],
        [InlineKeyboardButton("💰 Pay fine | Сплатити штраф", url="https://t.me/ShtrafyPDRbot")],
        [InlineKeyboardButton("Contacts | Контакти", callback_data="contacts")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    text = ("🇬🇧 EN\n"
            "Hello! This is the NPA Fleet bot 🚗\n"
            "I can help you create reports for vehicles.\n\n"
            "🇺🇦 UA\n"
            "Привіт! Це бот NPA Fleet 🚗\n"
            "Я допоможу вам створювати звіти по автомобілях.\n\n"
            "What are you interested in today? / Що вас цікавить сьогодні?"
            )
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.message.delete()
        except: pass
        await update.callback_query.message.reply_text(text=text, reply_markup=reply_markup)
    else:
        await update.message.reply_text(text=text, reply_markup=reply_markup)





# если нужно ограничить доступ — ставишь True
RESTRICTED_MODE = True

# =================== Работа с JSON ===================
USERS_JSON = "allowed_users.json"

def load_allowed_users():
    if os.path.exists(USERS_JSON):
        with open(USERS_JSON, "r", encoding="utf-8") as f:
            return {int(k): v for k, v in json.load(f).items()}
    return {}

def save_allowed_users():
    with open(USERS_JSON, "w", encoding="utf-8") as f:
        json.dump({str(k): v for k, v in ALLOWED_USERS.items()}, f, ensure_ascii=False, indent=4)

ALLOWED_USERS = load_allowed_users()





# =================== Ограничение доступа ===================
def restricted(func):
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        user_id = update.effective_user.id
        if user_id not in ALLOWED_USERS:
            if update.message:
                await update.message.reply_text("⛔ Доступ заборонений")
            elif update.callback_query:
                await update.callback_query.answer()
                await update.callback_query.message.reply_text("⛔ Доступ заборонений")
            return
        return await func(update, context, *args, **kwargs)
    return wrapper


# =================== Команды добавления/удаления пользователей ===================
@restricted
async def add_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id != ADMIN_ID:
        await update.message.reply_text("⛔ Лише адміністратор може додавати користувачів.")
        return

    args = context.args
    if len(args) < 2:
        await update.message.reply_text("Використання: /add_user <tg_id> <ім'я>")
        return

    try:
        new_id = int(args[0])
        name = " ".join(args[1:])
        ALLOWED_USERS[new_id] = name
        save_allowed_users()
        await update.message.reply_text(f"✅ Користувача {name} ({new_id}) додано до списку дозволених")
    except ValueError:
        await update.message.reply_text("⛔ Невірний ID")

@restricted
async def remove_user(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id != ADMIN_ID:
        await update.message.reply_text("⛔ Лише адміністратор може видаляти користувачів.")
        return

    args = context.args
    if len(args) != 1:
        await update.message.reply_text("Використання: /remove_user <tg_id>")
        return

    try:
        del_id = int(args[0])
        if del_id in ALLOWED_USERS:
            name = ALLOWED_USERS.pop(del_id)
            save_allowed_users()
            await update.message.reply_text(f"✅ Користувача {name} ({del_id}) видалено")
        else:
            await update.message.reply_text("⛔ Користувача не знайдено")
    except ValueError:
        await update.message.reply_text("⛔ Невірний ID")

# =======================================================================================================



#=================================Список пользователей на екран ТГ бота==================================
@restricted
async def list_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id != ADMIN_ID:
        await update.message.reply_text("⛔ Лише адміністратор може переглядати список користувачів")
        return

    if not ALLOWED_USERS:
        await update.message.reply_text("Список користувачів порожній.")
        return

    text = "📋 Список дозволених користувачів:\n\n"
    for uid, name in ALLOWED_USERS.items():
        text += f"- {name} ({uid})\n"

    await update.message.reply_text(text)
# =======================================================================================================




@restricted
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if RESTRICTED_MODE:
        if user_id not in ALLOWED_USERS:
            if update.message:
                await update.message.reply_text("⛔ Доступ заборонений")
            elif update.callback_query:
                await update.callback_query.answer()
                await update.callback_query.message.reply_text("⛔ Доступ заборонений")
            return  # прекращаем выполнение
        


    # очищаем user_data
    context.user_data.clear()

    # # готовим фото
    logo_bytes = get_logo_bytes()
    logo_file = InputFile(logo_bytes, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # отправляем фото
    if update.message:
        await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

@restricted
async def start_button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await main_menu(update, context)

# =================== Cancel ===================
@restricted
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.message.delete()
        except: pass
    await main_menu(update, context)
    return ConversationHandler.END


MANAGERS = {
    "Shyroke": [ADMIN_ID],
    "Mykolaiv": [6093640376, 6488832046],
    "Kyiv": [ADMIN_ID],     
    "Sumy/Romny": [ADMIN_ID]
}



#==========================================================================DAMAGE================================================================
# Інструкція при ДТП - головний пункт
@restricted
async def accident_procedure_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    text = "🌐 Оберіть мову / Choose language:"

    keyboard = [
        [InlineKeyboardButton("🇺🇦 Українська", callback_data="accident_procedure_ua")],
        [InlineKeyboardButton("🇬🇧 English", callback_data="accident_procedure_en")],
        [InlineKeyboardButton("⬅️ В головне меню", callback_data="main_menu")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        await query.message.delete()
    except:
        pass

    await query.message.reply_text(text=text, reply_markup=reply_markup)


# Інструкція при ДТП - українська
@restricted
async def accident_procedure_ua_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    text = (
    "🚨 **Порядок дій при ДТП** 🚨\n\n"
    "1️⃣ Негайно зверніться до місцевих органів влади, офісу NPA та співробітника автопарку, який контролює вашу подорож.\n"
    "2️⃣ Убезпечте місце аварії: виставте попереджувальний трикутник та попереджайте інший транспорт.\n"
    "3️⃣ Забезпечте комфорт потерпілим, не переміщуйте їх без необхідності.\n"
    "4️⃣ У разі потреби організуйте доставку поранених до лікарні.\n"
    "5️⃣ Не переміщуйте транспортний засіб до прибуття поліції.\n"
    "6️⃣ Не визнавайте жодної відповідальності.\n"    
    "7️⃣ Не залишайте місце ДТП.\n"
    "8️⃣ Фіксуйте подію (фото/відео), якщо це безпечно.\n"
    "9️⃣ Захистіть особисті речі.\n"
    "🔟 Уникайте суперечок з іншими учасниками.\n"
    "1️⃣1️⃣ Візьміть контакти та дані страхування інших учасників.\n"
    "1️⃣2️⃣ Запишіть марку, модель, колір та номерні знаки інших авто.\n"
    "1️⃣3️⃣ Підготуйте документи для перевірки поліцією (якщо потрібно, супроводжуйте поліцію до відділку)\n"
    "1️⃣4️⃣ Отримайте копію протоколу.\n"
    "1️⃣5️⃣ Повідомте офіцера автопарку, офіцера безпеки та керівника про подію.\n"
    "1️⃣6️⃣ Заповніть звіт про ДТП (VAR – додаток E) протягом 24 годин.\n"
)

    keyboard = [
        [InlineKeyboardButton("⬅️ Назад", callback_data="accident_procedure")],
        [InlineKeyboardButton("⬅️ В головне меню", callback_data="main_menu")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        await query.message.delete()
    except:
        pass

    await query.message.reply_text(text=text, reply_markup=reply_markup, parse_mode="Markdown")


# Інструкція при ДТП - англійська
@restricted
async def accident_procedure_en_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    text = (
    "🚨 **Accident Procedure** 🚨\n\n"
    "1️⃣ Immediately contact local authorities, the NPA office, and the fleet officer supervising your journey.\n"
    "2️⃣ Secure the accident site: place a warning triangle and alert approaching traffic.\n"
    "3️⃣ Ensure comfort for the injured, do not move them unless necessary.\n"
    "4️⃣ If needed, arrange for the injured to be transported to the hospital.\n"
    "5️⃣ Do not move the vehicle until the police arrive.\n"
    "6️⃣ Do not admit any liability.\n"
    "7️⃣ Do not leave the accident site.\n"
    "8️⃣ Document the incident (photos/videos) if safe.\n"
    "9️⃣ Protect your personal belongings.\n"
    "🔟 Avoid arguments or confrontations with other parties.\n"
    "1️⃣1️⃣ Obtain contact and insurance details from other parties involved.\n"
    "1️⃣2️⃣ Record the make, model, color, and registration numbers of other vehicles.\n"
    "1️⃣3️⃣ Prepare documents for police inspection (if required, accompany the police to the station).\n"
    "1️⃣4️⃣ Obtain a copy of the police report.\n"
    "1️⃣5️⃣ Inform the fleet officer, safety officer, and your project manager about the incident.\n"
    "1️⃣6️⃣ Complete the Vehicle Accident Report (VAR – Annex E) within 24 hours.\n"
)


    keyboard = [
        [InlineKeyboardButton("⬅️ Back", callback_data="accident_procedure")],
        [InlineKeyboardButton("⬅️ Main Menu", callback_data="main_menu")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)

    try:
        await query.message.delete()
    except:
        pass

    await query.message.reply_text(text=text, reply_markup=reply_markup, parse_mode="Markdown")



#=======================================================================END DAMAGE=============================================================================================









# ================================================================== LDR ========================================================================================================
SERIAL_LDR, ODOMETER_LDR, ALLOCATION_LDR, TEAM_NUMBER_LDR, USER_LDR, DESCRIPTION_LDR, OTHER_REQUEST_INPUT = range(7)


@restricted
async def ldr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    keyboard = [
        [InlineKeyboardButton("Flat tire | Проколоте колесо", callback_data="flat_tire")],
        [InlineKeyboardButton("Other damage | Інше пошкодження", callback_data="other_request")],
        [InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]
    ]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text("Choose request type:\nВиберіть тип звернення:", reply_markup=InlineKeyboardMarkup(keyboard))
    


@restricted
async def ldr_request_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "cancel":
        return await cancel(update, context)

    # Подготовка workbook
    wb = get_workbook("LDR")
    context.user_data['LDR'] = {
        'wb': wb,
        'ws': wb.active
    }
    ws = context.user_data['LDR']['ws']

    if data == "other_request":
        try:
            await query.message.delete()
        except:
            pass
        # Простейший ввод текста без удаления предыдущих сообщений
        await query.message.reply_text(
            "Please indicate what is damaged:\nВкажіть, що пошкоджено:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel", callback_data="cancel")]])
        )
        return OTHER_REQUEST_INPUT

    # Стандартные варианты
    if data == "flat_tire":
        set_cell(ws, "C5", "Flat tyre")
    elif data == "wipers":
        set_cell(ws, "C5", "Wipers replacement")
    elif data == "Drivers_card":
        set_cell(ws, "C5", "Driver's card")    

    set_cell(ws, "F5", "Serial / ID / Серійний номер / ID")

    # **Быстрый выбор локации**
    keyboard = [
        [InlineKeyboardButton(x, callback_data=x)] for x in ["Shyroke", "Mykolaiv", "Kyiv", "Sumy/Romny"]
    ]
    keyboard.append([InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")])
    try:
        await query.message.delete()
    except:
        pass
    await query.message.reply_text(
        "Select vehicle location:\nОберіть локацію автомобіля:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION_LDR




# Новый хэндлер для ввода текста пользователем
from googletrans import Translator

translator = Translator()

async def translate_to_en(text: str) -> str:
    # Асинхронно вызываем перевод
    translated = await translator.translate(text, dest='en')
    return translated.text

@restricted
async def ldr_other_request_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_text = update.message.text.strip()
    if not user_text:
        await update.message.reply_text("❌ Please type your request / ❌ Введіть ваше звернення")
        return OTHER_REQUEST_INPUT

    ws = context.user_data['LDR']['ws']

    # Перевод на английский
    translated_text = await translate_to_en(user_text)

    # Записываем перевод в Excel
    set_cell(ws, "C5", translated_text)
    set_cell(ws, "F5", "Serial / ID / Серійний номер / ID")

    


    keyboard = [
        [InlineKeyboardButton("Shyroke", callback_data="Shyroke")],
        [InlineKeyboardButton("Mykolaiv", callback_data="Mykolaiv")],
        [InlineKeyboardButton("Kyiv", callback_data="Kyiv")],
        [InlineKeyboardButton("Sumy/Romny", callback_data="Sumy/Romny")],
        [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
    ]
    await update.message.reply_text(
        "Select vehicle location:\nОберіть локацію автомобіля:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION_LDR

# =================== Ввод данных ===================


async def serial_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().upper()
    text = text.replace(" ", "")

    if re.fullmatch(r"[A-Z]{2}\d{2}", text):
        text = text[:2] + "-" + text[2:]

    if not re.fullmatch(r"[A-Z]{2}-\d{2}", text):
        await update.message.reply_text(
            "❌ Формат повинен бути:(напр. HP-01)\n        Format must be:(e.g. HP-01)"
        )
        return SERIAL_LDR

    ws = context.user_data['LDR']['ws']
    set_cell(ws, "F5", text)

    

    await update.message.reply_text(
        "Enter current odometer value (km):\nВведіть поточний пробіг (км):",
        reply_markup=InlineKeyboardMarkup(
            [[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]]
        )
    )
    return ODOMETER_LDR




async def odometer_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    if not text.isdigit():
        await update.message.reply_text("❌ Odometer must be a number (in km)\n❌ Пробіг повинен бути числом (в км)")
        return ODOMETER_LDR

    ws = context.user_data['LDR']['ws']
    set_cell(ws, "I8", int(text))  # например пишем пробег в C9

    

    # После одометра → выбор Allocation
    keyboard = [
        [InlineKeyboardButton(x, callback_data=x)] for x in ["MTT","MDD","MECH","NTS","OPS/SUPP","ADMIN"]
    ]
    keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])

    await update.message.reply_text(
        "Choose Allocation:\nОберіть Розподіл:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION_LDR



async def allocation_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    selection = query.data

    # Проверяем наличие workbook
    if 'LDR' not in context.user_data or 'ws' not in context.user_data['LDR']:
        await query.message.reply_text(
            "❌ Please start the request from the beginning using /start\n❌ Будь ласка, почніть звернення заново за допомогою /start"
        )
        return ConversationHandler.END

    ws = context.user_data['LDR']['ws']

    # Обработка отмены
    if selection == "cancel":
        return await cancel(update, context)

    # Локации Shyroke / Mykolaiv
    if selection in ["Shyroke", "Mykolaiv", "Kyiv", "Sumy/Romny"]:
        context.user_data['location'] = selection
        set_cell(ws, "C8", selection)
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter vehicle call sign (e.g. HP-01): \nВведіть внутрішній номер авто (напр. HP-01):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return SERIAL_LDR

    # Если пользователь выбрал OPS/SUPP — показываем второй уровень кнопок
    if selection == "OPS/SUPP":
        keyboard = [[InlineKeyboardButton(x, callback_data=f"OPS/{x}")] for x in ["STFM","TFM","SUPV","LOGS","IMM","QA"]]
        keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Choose sub-allocation for OPS/SUPP:\nОберіть підрозподіл для OPS/SUPP:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ALLOCATION_LDR  # остаёмся на этом же шаге, ждём второй выбор

    # Обработка выбора подкнопки OPS/SUPP
    if selection.startswith("OPS/"):
        allocation_choice = selection.split("/")[1]
        set_cell(ws, "F8", f"{allocation_choice}")

        # ✅ сохраняем в namespace LDR
        context.user_data['LDR']['allocation'] = allocation_choice

        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter your full name:\nВведіть ваше Ім'я та прізвище:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return USER_LDR


    # Если MTT, MDD, NTS — спрашиваем номер команды
    if selection.upper() in ["MTT", "MDD", "NTS"]:
        context.user_data['LDR']['allocation'] = selection.upper()

        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            f"Enter team number for {selection.upper()}:\nВведіть номер команди для {selection.upper()}:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return TEAM_NUMBER_LDR

    # Если MECH — просто записываем в Excel и спрашиваем имя
    if selection.upper() in ("MECH", "ADMIN"):
        set_cell(ws, "F8", selection.upper())
        try: 
            await query.message.delete()
        except: 
            pass
        await query.message.reply_text(
            "Enter your full name:\nВведіть ваше Ім'я та прізвище:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return USER_LDR


async def team_number_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ Team number must be a number")
        return TEAM_NUMBER_LDR

    ws = context.user_data['LDR']['ws']
    allocation = context.user_data['LDR'].get('allocation', '')
    set_cell(ws, "F8", f"{allocation}-{text}")

    

    await update.message.reply_text(
        "Enter your full name:\nВведіть Ім'я та прізвище:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
    )
    return USER_LDR



async def user_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ You did not enter your name")
        return USER_LDR
    user_name_latin = unidecode(text)
    ws = context.user_data['LDR']['ws']
    set_cell(ws, "I5", user_name_latin)
    set_cell(ws, "B19", user_name_latin)
    location = context.user_data.get('location')
    manager_fa = {"Shyroke": "F.A. Oleksandr Rudnov",
                  "Mykolaiv": "F.A. Andriy Padalka",
                  "Kyiv": "F.A. Oleksandr Rudnov",
                  "Sumy/Romny": "F.A. Oleksandr Rudnov"}.get(location,"F.A. Unknown")
    set_cell(ws, "F19", manager_fa)
    set_cell(ws, "C19", datetime.now().strftime("%d-%m-%Y"))

    

    await update.message.reply_text(
        "Detailed description of events leading to the loss or damage:\nДетальний опис подій, що призвели до втрати або пошкодження:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]])
    )
    return DESCRIPTION_LDR



# =================== Описание ===================




def auto_height_for_cell(ws, cell_address):
    cell = ws[cell_address]
    cell.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=True)

    # Получаем ширину колонки в символах (приближённо)
    col_letter = ''.join(filter(str.isalpha, cell_address))
    col_width = ws.column_dimensions[col_letter].width or 10  # если не задано, ставим 10

    # Оценка количества строк: длина текста / ширина колонки
    text_length = len(str(cell.value))
    lines_needed = math.ceil(text_length / col_width)

    # Стандартная высота одной строки ~15
    ws.row_dimensions[cell.row].height = lines_needed * 15

# Пример использования в твоей функции:










def split_text(text, words_per_line=12):
    """Разбивает текст на строки примерно по 20 слов"""
    words = text.split()
    return [" ".join(words[i:i+words_per_line]) for i in range(0, len(words), words_per_line)]

async def description_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ Describe the situation:\n❌ Опишіть ситуацію")
        return DESCRIPTION_LDR

    text_en = await translate_to_en(text)
    ws = context.user_data['LDR']['ws']

    # Разбиваем текст на куски
    lines = split_text(text_en, words_per_line=20)

    # вставка текста по строкам
    start_row = 13  # теперь B13
    for i, line in enumerate(lines, start=start_row):
        if i > 20:
            break
        cell = ws[f"B{i}"]
        cell.value = line
        cell.alignment = Alignment(horizontal="left", vertical="bottom")


    # Подгоняем размеры остальных ячеек
    auto_adjust(ws, ["C5","F5","C8","F8","I5","B19","C19","F19"])

    
    

    plate = ws["F5"].value or "CAR"
    filename = f"LDR_{plate}_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"

    # Отправка менеджерам по локации


    # Отправка менеджерам по локации
    location = context.user_data.get("location")
    manager_ids = MANAGERS.get(location, [])
    user_id = update.effective_user.id
    user_name = ALLOWED_USERS.get(user_id, "Unknown")  # получаем имя из словаря

    for manager_id in manager_ids:
        file_stream = BytesIO()
        ws.parent.save(file_stream)
        file_stream.seek(0)
        await context.bot.send_document(chat_id=manager_id, document=file_stream, filename=filename)
        await context.bot.send_message(
            chat_id=manager_id,
            text=f"📄 Новий LDR звіт по локації {location} від {user_name}"
        )

    context.user_data.clear()

    # Уведомление пользователю
    await update.message.reply_text("✅ Звіт надіслано Fleet співробітнику, відповідно до обраної локації.\nВам залишилось лише підписати його.\n\n✅ The report has been sent to the Fleet of chosen location.\n You only need to sign it.")
    
    
    # Задержка 3 секунды
    await asyncio.sleep(5)


    # Приветственное фото с кнопкой
    logo_bytes_start = get_logo_bytes()
    logo_file = InputFile(logo_bytes_start, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    try: 
        await update.message.delete()
    except: 
        pass

    await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

    return ConversationHandler.END

# Функция для авто-подгонки высоты строки A9 с минимальной защитой
def auto_height_for_cell(ws, cell_address, min_height=45):
    """Автоматическая высота строки под содержимое, но не меньше min_height"""
    cell = ws[cell_address]
    row = cell.row
    lines = str(cell.value).count('\n') + 1
    # Расчет высоты: 15 пикселей на строку
    height = max(lines * 15, min_height)
    ws.row_dimensions[row].height = height






# =================== Заглушки ===================
async def generic_stub(update: Update, context: ContextTypes.DEFAULT_TYPE, name="Function"):
    query = update.callback_query
    await query.answer()
    keyboard = [[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(f"You selected {name}. Function in progress.", reply_markup=InlineKeyboardMarkup(keyboard))


#=====================================================LDR END=============================================================================
















    #=====================================================MFR=================================================================================
# Пример загрузки JSON при старте бота
import json

with open("asset_num.json", "r", encoding="utf-8") as f:
    ASSET_NUMBERS = json.load(f)




    # ------------------------- Константы состояний -------------------------
    MFR_ALLOCATION, MFR_MODEL_SELECTION, MFR_SERIAL, MFR_ODOMETER, MFR_TEAM_NUMBER, MFR_USER, MFR_DESCRIPTION = range(7)



    # ------------------------- Начало MFR -------------------------
    @restricted
    async def mfr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()

        # ================= MFR namespace =================
        if 'MFR' not in context.user_data:
            context.user_data['MFR'] = {}
        mfr_data = context.user_data['MFR']

        mfr_data['wb'] = get_workbook("MFR")
        mfr_data['ws'] = mfr_data['wb'].active
        ws = mfr_data['ws']

        set_cell(ws, "F6", "Serial / ID / Серійний номер / ID")

        keyboard = [
            [InlineKeyboardButton("Shyroke", callback_data="Shyroke")],
            [InlineKeyboardButton("Mykolaiv", callback_data="Mykolaiv")],
            [InlineKeyboardButton("Kyiv", callback_data="Kyiv")],
            [InlineKeyboardButton("Sumy/Romny", callback_data="Sumy/Romny")],
            [InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]
        ]

        try: await query.message.delete()
        except: pass

        await query.message.reply_text(
            "Select vehicle location:\nОберіть локацію автомобіля:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return MFR_ALLOCATION

    # ------------------------- Выбор локации -------------------------
    async def mfr_location_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        location = query.data
        if location == "cancel":
            return await cancel(update, context)

        ws = context.user_data['MFR']['ws']
        set_cell(ws, "C9", location)
        context.user_data['location'] = location

        try: await query.message.delete()
        except: pass

        keyboard = [
            [InlineKeyboardButton("TOYOTA", callback_data="brand_TOYOTA")],
            [InlineKeyboardButton("FORD", callback_data="brand_FORD")],
            [InlineKeyboardButton("MITSUBISHI", callback_data="brand_MITSUBISHI")],
            [InlineKeyboardButton("VOLKSWAGEN", callback_data="brand_VOLKSWAGEN")],
            [InlineKeyboardButton("RENAULT DUSTER", callback_data="RENAULT DUSTER")],
            [InlineKeyboardButton("SKODA KODIAQ", callback_data="SKODA KODIAQ")],
            [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
        ]

        await query.message.reply_text(
            "Select car brand:\nОберіть марку авто:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return MFR_MODEL_SELECTION

    # ------------------------- Выбор модели авто -------------------------
    async def model_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        choice = query.data
        ws = context.user_data['MFR']['ws']

        if choice == "cancel":
            return await cancel(update, context)

        if choice.startswith("brand_"):
            brand = choice.replace("brand_", "")
            if brand == "TOYOTA": models = ["Toyota Hilux", "Toyota Land Cruiser"]
            elif brand == "FORD": models = ["Ford Ranger", "Ford Transit", "Ford Truck"]
            elif brand == "MITSUBISHI": models = ["Mitsubishi L200", "Mitsubishi ASX", "Mitsubishi Outlander"]
            elif brand == "VOLKSWAGEN": models = ["Volkswagen T6", "Volkswagen ID.4"]
            else: models = []

            keyboard = [[InlineKeyboardButton(m, callback_data=m)] for m in models]
            keyboard.append([InlineKeyboardButton("⬅️ Back | Назад", callback_data="back_to_brands")])
            keyboard.append([InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")])

            try: await query.message.delete()
            except: pass

            await query.message.reply_text(
                f"Select model of {brand}:\nВиберіть модель {brand}:",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
            return MFR_MODEL_SELECTION

        if choice == "back_to_brands":
            return await mfr_location_selection(update, context)

        # Если выбрана конкретная модель
        set_cell(ws, "C6", choice)

        try: await query.message.delete()
        except: pass

        await query.message.reply_text(
            "Enter vehicle call sign (e.g. HP-01):\nВведіть внутрішній номер авто (напр. HP-01):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return MFR_SERIAL

    # ------------------------- Ввод номера авто -------------------------
    async def serial_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
        text = update.message.text.strip().upper().replace(" ", "")
        
        # Приводим к формату HP-01
        if re.fullmatch(r"[A-Z]{2}\d{2}", text):
            text = text[:2] + "-" + text[2:]

        # Проверка формата
        if not re.fullmatch(r"[A-Z]{2}-\d{2}", text):
            await update.message.reply_text(
                "❌ Формат повинен бути:(напр. HP-01)\nFormat must be:(e.g. HP-01)"
            )
            return MFR_SERIAL

        # ========================= Используем отдельный namespace для MFR =========================
        if 'MFR' not in context.user_data:
            context.user_data['MFR'] = {}
            
        mfr_data = context.user_data['MFR']

        if 'ws' not in mfr_data:
            mfr_data['wb'] = get_workbook("MFR")
            mfr_data['ws'] = mfr_data['wb'].active

        ws = mfr_data['ws']

        # Получаем asset_number из JSON
        asset_data = ASSET_NUMBERS.get(text)
        asset_number = asset_data['asset_number'] if asset_data else "UNKNOWN"

        # Записываем в Excel
        set_cell(ws, "C11", asset_number)
        set_cell(ws, "F5", text)  # Внутренний номер авто

        await update.message.reply_text(
            "Enter odometer reading (km):\nВведіть поточний пробіг (км):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )

        return MFR_ODOMETER




    # ------------------------- Ввод одометра -------------------------
    async def odometer_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
        text = update.message.text.strip()
        if not text.isdigit():
            await update.message.reply_text("❌ Odometer must be a number\n❌ Одометр повинен бути числом")
            return MFR_ODOMETER

        ws = context.user_data['MFR']['ws']
        set_cell(ws, "I8", text)

        keyboard = [[InlineKeyboardButton(x, callback_data=x)] for x in ["MTT", "MDD", "MECH", "NTS", "OPS/SUPP", "ADMIN"]]
        keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])

        await update.message.reply_text(
            "Choose Allocation:\nОберіть підрозділ:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return MFR_ALLOCATION



    @restricted
    async def allocation_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
        query = update.callback_query
        await query.answer()
        ws = context.user_data['MFR']['ws']
        selection = query.data

        if selection == "cancel":
            return await cancel(update, context)

        # ---------------- Локации ----------------
        if selection in ["Shyroke", "Mykolaiv", "Kyiv", "Sumy/Romny"]:
            context.user_data['location'] = selection
            set_cell(ws, "C9", selection)
            try:
                await query.message.delete()
            except:
                pass
            await query.message.reply_text(
                "Enter vehicle call sign (e.g. HP-01):",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]]
                )
            )
            return MFR_SERIAL

        # ---------------- OPS/SUPP ----------------
        if selection == "OPS/SUPP":
            keyboard = [[InlineKeyboardButton(x, callback_data=f"OPS/{x}")]
                        for x in ["STFM", "TFM", "SUPV", "LOGS", "IMM", "QA"]]
            keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])

            try:
                await query.message.edit_text(
                    "Choose sub-allocation for OPS/SUPP:\nОберіть підрозподіл для OPS/SUPP:",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
            except:
                await query.message.reply_text(
                    "Choose sub-allocation for OPS/SUPP:\nОберіть підрозподіл для OPS/SUPP:",
                    reply_markup=InlineKeyboardMarkup(keyboard)
                )
            return MFR_ALLOCATION

        # ---------------- OPS/подразделения ----------------
        if selection.startswith("OPS/"):
            allocation_choice = selection.split("/")[1]
            set_cell(ws, "F9", f"{allocation_choice}")
            try:
                await query.message.delete()
            except:
                pass
            await query.message.reply_text(
                "Enter your full name:\nВведіть ваше ім'я та прізвище:",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]]
                )
            )
            return MFR_USER

        # ---------------- MTT/MDD/NTS ----------------
        if selection.upper() in ["MTT", "MDD", "NTS"]:
            context.user_data['allocation'] = selection.upper()
            try:
                await query.message.delete()
            except:
                pass
            await query.message.reply_text(
                f"Enter team number for {selection.upper()}:\nВведіть номер команди для {selection.upper()}:",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]]
                )
            )
            return MFR_TEAM_NUMBER

        # ---------------- MECH/ADMIN ----------------
        if selection.upper() in ["MECH", "ADMIN"]:
            set_cell(ws, "F8", selection.upper())
            try:
                await query.message.delete()
            except:
                pass
            await query.message.reply_text(
                "Enter your full name:\nВведіть ваше ім'я та прізвище:",
                reply_markup=InlineKeyboardMarkup(
                    [[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]]
                )
            )
            return MFR_USER






    # ------------------------- Team Number -------------------------
    async def team_number_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
        text = update.message.text.strip()
        if not text.isdigit():
            await update.message.reply_text("❌ Team number must be a number")
            return MFR_TEAM_NUMBER

        ws = context.user_data['MFR']['ws']
        allocation = context.user_data.get('allocation')
        set_cell(ws, "F9", f"{allocation}-{text}")

        await update.message.reply_text(
            "Enter your full name:\nВведіть ім'я та прізвище:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return MFR_USER

    # ------------------------- User -------------------------
    async def user_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
        text = update.message.text.strip()
        if not text:
            await update.message.reply_text("❌ You did not enter your name")
            return MFR_USER

        # Преобразуем в латиницу и делаем каждое слово с заглавной буквы
        user_name_latin = unidecode(text).title()

        ws = context.user_data['MFR']['ws']
        set_cell(ws, "I6", user_name_latin)
        set_cell(ws, "B22", user_name_latin)

        location = context.user_data.get('location')
        manager_fa = {
            "Shyroke": "F.A. Oleksandr Rudnov",
            "Mykolaiv": "F.A. Andriy Padalka",
            "Kyiv": "F.A. Oleksandr Rudnov",
            "Sumy/Romny": "F.A. Oleksandr Rudnov"
        }.get(location, "F.A. Unknown")
        set_cell(ws, "F22", manager_fa)
        set_cell(ws, "C22", datetime.now().strftime("%d-%m-%Y"))
        set_cell(ws, "F12", datetime.now().strftime("%d-%m-%Y"))

        # Сообщение пользователю для следующего шага
        await update.message.reply_text(
            "Please describe the mechanical issue in detail:\nБудь ласка, опишіть механічну несправність детально:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
        )
        return MFR_DESCRIPTION


    # ------------------------- Description -------------------------
    def split_text(text, words_per_line=20):
        words = text.split()
        return [" ".join(words[i:i+words_per_line]) for i in range(0, len(words), words_per_line)]

    async def description_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
        text = update.message.text.strip()
        if not text:
            await update.message.reply_text("❌ Describe the situation:\n❌ Опишіть ситуацію")
            return MFR_DESCRIPTION

        text_en = await translate_to_en(text)
        ws = context.user_data['MFR']['ws']

        lines = split_text(text_en, words_per_line=20)
        start_row = 16
        for i, line in enumerate(lines, start=start_row):
            if i > 21: break
            ws[f"B{i}"].value = line
            ws[f"B{i}"].alignment = Alignment(horizontal="left", vertical="bottom")

        auto_adjust(ws, ["F5","C6","C9","F9","I6","F22","C22"])

        plate = ws["F5"].value or "CAR"
        filename = f"MFR_{plate}_{datetime.now().strftime('%d-%m-%Y_%H-%M')}.xlsx"

        location = context.user_data.get("location")
        manager_ids = MANAGERS.get(location, [])
        user_id = update.effective_user.id
        user_name = ALLOWED_USERS.get(user_id,"Unknown")

        for manager_id in manager_ids:
            file_stream = BytesIO()
            ws.parent.save(file_stream)
            file_stream.seek(0)
            await context.bot.send_document(chat_id=manager_id, document=file_stream, filename=filename)
            await context.bot.send_message(
                chat_id=manager_id,
                text=f"📄 Новий MFR звіт по локації {location} від {user_name}"
            )

        context.user_data.clear()
        await update.message.reply_text(
            "✅ Звіт надіслано Fleet співробітнику, відповідно до обраної локації.\nВам залишилось лише підписати його.\n\n✅ The report has been sent to the Fleet of chosen location.\n You only need to sign it."
        )


        # Задержка 3 секунды
        await asyncio.sleep(5)


        logo_bytes_start = get_logo_bytes()
        logo_file = InputFile(logo_bytes_start, filename="logo.png")
        keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

        return ConversationHandler.END



#=============================================================MFR END=============================================================















#=============================================================Monthly inspection form=============================================================

import os
import re
import shutil
import asyncio
from telegram import InputFile
from datetime import datetime
from openpyxl import load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ConversationHandler, MessageHandler, CallbackQueryHandler,
    filters, ContextTypes
)
from monthly_questions import MONTHLY_QUESTIONS
import json
import xlwings as xw

# =================== Константы ===================
ADMIN_ID = 507775858
EXCEL_TEMPLATE = os.path.join("excel", "MIF.xlsx")
RESULT_FOLDER = "Result"
os.makedirs(RESULT_FOLDER, exist_ok=True)

# =================== FSM состояния ===================
(SELECT_LOCATION, SELECT_BRAND, CALLSIGN, ODOMETER, USER, QUESTION, REASON, PHOTO) = range(8)

# =================== Менеджеры ===================
MANAGERS = {
    "Shyroke": [ADMIN_ID],
    "Mykolaiv": [6093640376, 6488832046],
    "Kyiv": [ADMIN_ID],
    "Sumy/Romny": [ADMIN_ID]
}

# =================== Вопросы ===================
MONTHLY_QUESTIONS = MONTHLY_QUESTIONS

# =================== Call sign → Registration ===================
with open("cars.json", "r", encoding="utf-8") as f:
    CARS = json.load(f)

# =================== Функции для Excel ===================
def set_cell(ws, cell, value):
    try:
        ws[cell].value = value
    except AttributeError:
        for merged_range in ws.merged_cells.ranges:
            if cell in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col, value=value)
                break

def save_all_to_excel(user_data, folder_path, excel_filename):
    """
    Сохраняет данные пользователя в Excel-шаблон через xlwings,
    полностью сохраняя шрифты, стили и объединенные ячейки.
    """
    file_path = os.path.join(folder_path, excel_filename)
    os.makedirs(folder_path, exist_ok=True)

    # 1️⃣ Копируем шаблон
    template_path = os.path.join("excel", "MIF.xlsx")
    shutil.copyfile(template_path, file_path)

    # 2️⃣ Запускаем Excel в фоновом режиме
    app = xw.App(visible=False)
    try:
        wb = xw.Book(file_path)
        ws = wb.sheets[0]

        # ======= Общие данные =======
        ws.range("A4").value = datetime.now().strftime("%Y-%m-%d")
        ws.range("B4").value = user_data.get("brand", "")
        ws.range("C4").value = user_data.get("registration_number", "")
        ws.range("E4").value = user_data.get("call_sign", "")
        ws.range("G4").value = user_data.get("odometer", "")
        ws.range("H4").value = user_data.get("driver_name", "")

        # ======= Вопросы =======
        for idx, q_data in enumerate(MONTHLY_QUESTIONS):
            ans = user_data['answers'].get(idx, {})
            if ans.get('yes'):
                ws.range(q_data['yes_cell']).value = "+"  # сохраняем только значение
            if ans.get('no'):
                ws.range(q_data['no_cell']).value = "-"
                ws.range(q_data['remark_cell']).value = ans.get('remark', '')

        # Сохраняем и закрываем
        wb.save()
        wb.close()
    finally:
        app.quit()

    return file_path

# =================== START ===================
async def start_inspection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("Shyroke", callback_data="loc_Shyroke")],
        [InlineKeyboardButton("Mykolaiv", callback_data="loc_Mykolaiv")],
        [InlineKeyboardButton("Kyiv", callback_data="loc_Kyiv")],
        [InlineKeyboardButton("Sumy/Romny", callback_data="loc_Sumy/Romny")],
        [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
    ]
    msg = update.message or update.callback_query.message
    await msg.reply_text("Select location\nОберіть локацію:", reply_markup=InlineKeyboardMarkup(keyboard))
    return SELECT_LOCATION

# =================== LOCATION SELECT ===================
async def location_choice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        return await cancel(update, context)
    context.user_data['location'] = query.data.replace("loc_", "")
    keyboard = [
        [InlineKeyboardButton("TOYOTA", callback_data="brand_TOYOTA")],
        [InlineKeyboardButton("FORD", callback_data="brand_FORD")],
        [InlineKeyboardButton("MITSUBISHI", callback_data="brand_MITSUBISHI")],
        [InlineKeyboardButton("VOLKSWAGEN", callback_data="brand_VOLKSWAGEN")],
        [InlineKeyboardButton("RENAULT DUSTER", callback_data="brand_RENAULT DUSTER")],
        [InlineKeyboardButton("SKODA KODIAQ", callback_data="brand_SKODA KODIAQ")],
        [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
    ]
    await query.message.reply_text("Select car brand\nОберіть марку авто:", reply_markup=InlineKeyboardMarkup(keyboard))
    return SELECT_BRAND

# =================== BRAND SELECT ===================
async def brand_selected(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        return await cancel(update, context)
    context.user_data['brand'] = query.data.replace("brand_", "")
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]])
    await query.message.reply_text(
        "Enter call sign (HP-01)\nВведіть внутрішній номер авто:", reply_markup=keyboard
    )
    return CALLSIGN

# =================== CALLSIGN ===================
async def call_sign_input(update, context):
    text = update.message.text.strip().upper().replace(" ", "").replace("_", "")
    if text.lower() in ["cancel", "відмінити", "❌"]:
        return await cancel(update, context)

    match = re.fullmatch(r"([A-Z]{2})-?(\d{2})", text)
    if not match:
        keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]])
        await update.message.reply_text("❌ Формат повинен бути HP-01", reply_markup=keyboard)
        return CALLSIGN

    formatted_call_sign = f"{match[1]}-{match[2]}"
    registration_number = CARS.get(formatted_call_sign)
    if not registration_number:
        await update.message.reply_text(f"❌ Call sign {formatted_call_sign} не знайдено в базі.")
        return CALLSIGN

    context.user_data['call_sign'] = formatted_call_sign
    context.user_data['registration_number'] = registration_number

    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]])
    await update.message.reply_text(
        f"Call sign: {formatted_call_sign}\nРег. номер авто: {registration_number}\n\nEnter odometer reading (km)\nВведіть пробіг авто:",
        reply_markup=keyboard
    )
    return ODOMETER

# =================== ODOMETER ===================
async def odometer_input(update, context):
    text = update.message.text.strip()
    if text.lower() in ["cancel", "відмінити", "❌"]:
        return await cancel(update, context)
    if not text.isdigit():
        keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]])
        await update.message.reply_text("❌ Одометр повинен бути числом", reply_markup=keyboard)
        return ODOMETER
    context.user_data['odometer'] = text
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]])
    await update.message.reply_text("Enter your full name\nВведіть ваше ПІБ:", reply_markup=keyboard)
    return USER

# =================== USER ===================
async def user_input(update, context):
    text = update.message.text.strip()
    if text.lower() in ["cancel", "відмінити", "❌"]:
        return await cancel(update, context)
    context.user_data['driver_name'] = text
    context.user_data['answers'] = {}
    context.user_data['current_q'] = 0
    await ask_question(update, context)
    return QUESTION

# =================== QUESTIONS ===================
async def ask_question(update, context):
    idx = context.user_data['current_q']
    q = MONTHLY_QUESTIONS[idx]['text']
    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("Yes / Так", callback_data="yes")],
        [InlineKeyboardButton("No / Ні", callback_data="no")],
        [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
    ])
    msg = update.message or update.callback_query.message
    if update.callback_query:
        await update.callback_query.answer()
    await msg.reply_text(q, reply_markup=keyboard)

async def handle_question(update, context):
    query = update.callback_query
    await query.answer()
    if query.data == "cancel":
        return await cancel(update, context)
    idx = context.user_data['current_q']
    if query.data == "yes":
        context.user_data['answers'][idx] = {'yes': True}
        context.user_data['current_q'] += 1
        if context.user_data['current_q'] >= len(MONTHLY_QUESTIONS):
            return await finish_form(update, context)
        await ask_question(update, context)
        return QUESTION
    else:
        await query.message.reply_text("Enter reason\nВведіть зауваження:")
        return REASON

# =================== REASON ===================
async def handle_reason(update, context):
    text = update.message.text.strip()
    if text.lower() in ["cancel", "відмінити", "❌"]:
        return await cancel(update, context)
    context.user_data['reason'] = text
    keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]])
    await update.message.reply_text("Send photo\nНадішліть фото:", reply_markup=keyboard)
    return PHOTO

# =================== PHOTO ===================
async def handle_photo(update, context):
    idx = context.user_data['current_q']
    reason = context.user_data.pop('reason', '')

    location = context.user_data.get('location', 'UNKNOWN')
    call_sign = context.user_data.get('call_sign', 'UNKNOWN')
    folder_path = os.path.join("Result", location, call_sign)
    os.makedirs(folder_path, exist_ok=True)

    if update.message.photo:
        photo_file = await update.message.photo[-1].get_file()
        filename = f"photo_{update.effective_user.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.jpg"
        filepath = os.path.join(folder_path, filename)
        await photo_file.download_to_drive(filepath)
        full_remark = f"{reason}. \nФото: {filename}" if reason else f"Фото: {filename}"
    else:
        full_remark = reason

    context.user_data['answers'][idx] = {'no': True, 'remark': full_remark}
    context.user_data['current_q'] += 1

    if context.user_data['current_q'] >= len(MONTHLY_QUESTIONS):
        return await finish_form(update, context)
    await ask_question(update, context)
    return QUESTION

# =================== FINISH ===================
async def finish_form(update, context):
    location = context.user_data.get('location', 'UNKNOWN')
    call_sign = context.user_data.get('call_sign', 'UNKNOWN')
    folder_path = os.path.join("Result", location, call_sign)
    os.makedirs(folder_path, exist_ok=True)

    final_name = f"193-VMR-{datetime.now().strftime('%y-%b').upper()} {call_sign}"
    excel_filename = f"{final_name}.xlsx"

    file_path = save_all_to_excel(context.user_data, folder_path, excel_filename)
    manager_ids = MANAGERS.get(location, [])

    for manager_id in manager_ids:
        with open(file_path, "rb") as f:
            await context.bot.send_document(chat_id=manager_id, document=f, filename=excel_filename)
        await context.bot.send_message(chat_id=manager_id, text=f"📄 New report VMR for location {location}")

    context.user_data.clear()
    msg = update.message or (update.callback_query and update.callback_query.message)

    if msg:
        await msg.reply_text("✅ Report sent successfully!")
        await asyncio.sleep(2)
        logo_bytes_start = get_logo_bytes()
        logo_file = InputFile(logo_bytes_start, filename="logo.png")
        keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await msg.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

    return ConversationHandler.END

# =================== CANCEL ===================
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.message.delete()
        except: pass
    await main_menu(update, context)
    return ConversationHandler.END

# =================== HANDLER ===================
inspection_handler = ConversationHandler(
    entry_points=[CallbackQueryHandler(start_inspection, pattern="^monthly_form$")],
    states={
        SELECT_LOCATION: [CallbackQueryHandler(location_choice)],
        SELECT_BRAND: [CallbackQueryHandler(brand_selected)],
        CALLSIGN: [MessageHandler(filters.TEXT & ~filters.COMMAND, call_sign_input)],
        ODOMETER: [MessageHandler(filters.TEXT & ~filters.COMMAND, odometer_input)],
        USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input)],
        QUESTION: [CallbackQueryHandler(handle_question)],
        REASON: [MessageHandler(filters.TEXT & ~filters.COMMAND, handle_reason)],
        PHOTO: [MessageHandler(filters.PHOTO | (filters.TEXT & ~filters.COMMAND), handle_photo)],
    },
    fallbacks=[CallbackQueryHandler(cancel, pattern="^cancel$")]
)


#=============================================================END Monthly inspection form=============================================================






#===================================================================CONTACTS====================================================

@restricted
async def contacts_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Если нажата кнопка "Назад", возвращаемся в главное меню
    if query.data == "back":
        try:
            await query.message.delete()
        except:
            pass
        await main_menu(update, context)  # вызываем функцию главного меню
        return

    text = (
        "📌 Locations / Локації:\n"
        "Select a location to see contacts:\nОберіть локацію для контактів:"
    )

    keyboard = [
        [
            InlineKeyboardButton("Shyroke | Широке", callback_data="contact_shyroke"),
            InlineKeyboardButton("Mykolaiv | Миколаїв", callback_data="contact_mykolaiv"),
        ],
        [InlineKeyboardButton("⬅️ Back | Назад", callback_data="back")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)
    try:
        await query.message.delete()
    except:
        pass
    await query.message.reply_text(text=text, reply_markup=reply_markup)



#Обработчик конкретной локации
@restricted
async def contact_location_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "back":
        try: await query.message.delete()
        except: pass
        await main_menu(update, context)
        return

    if data == "contact_shyroke":
        text = (
            "📌 Shyroke | Широке\n"
            "👤 F.A. Oleksandr Rudnov | F.A. Олександр Руднов\n"
            "📞 Phone: +380 431 019 082\n"
            "🌐 Map: https://goo.gl/maps/example1"
        )
        keyboard = [
            [InlineKeyboardButton("Car Wash | Мийка", url="https://goo.gl/maps/carwash_shyroke")],
            [InlineKeyboardButton("Tire Service | Шиномонтаж", url="https://goo.gl/maps/tire_shyroke")],
            [InlineKeyboardButton("⬅️ Back | Назад", callback_data="contacts")]
        ]
    elif data == "contact_mykolaiv":
        text = (
            "📌 Mykolaiv | Миколаїв\n"
            "👤 F.A. Andriy Padalka | F.A. Андрій Падалка\n"
            "📞 Phone: +380 431 019 083\n"
            "🌐 Map: https://goo.gl/maps/example2"
        )
        keyboard = [
            [InlineKeyboardButton("Car Wash | Мийка", url="https://goo.gl/maps/carwash_mykolaiv")],
            [InlineKeyboardButton("Tire Service | Шиномонтаж", url="https://goo.gl/maps/tire_mykolaiv")],
            [InlineKeyboardButton("⬅️ Back | Назад", callback_data="contacts")]
        ]

    reply_markup = InlineKeyboardMarkup(keyboard)
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(text=text, reply_markup=reply_markup)


from locations import LOCATIONS

# пример использования
loc = "shyroke"
manager_name = LOCATIONS[loc]["manager"]["name"]
print(manager_name)  # => Oleksandr Rudnov | Олександр Руднов


@restricted
async def contact_location_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "back":
        try: await query.message.delete()
        except: pass
        await contacts_callback(update, context)
        return

    loc_key = None
    action = None

    if data.startswith("contact_"):
        loc_key = data.split("_")[1]  # shyroke или mykolaiv
        loc_data = LOCATIONS[loc_key]
        manager = loc_data["manager"]
        senior = loc_data["senior_officer"]

        text = (
            f"📌 {loc_key.capitalize()}\n\n"
            f"👤 Fleet Assistant: {manager['name']}\n"
            f"📞 Phone: {manager['phone']}\n"
            f"✉️ Email: {manager['email']}\n\n"
            f"👔 {senior['position']}: {senior['name']}\n"
            f"📞 Phone: {senior['phone']}\n"
            f"✉️ Email: {senior['email']}\n\n"  
        )

        keyboard = [
            [InlineKeyboardButton("🧼 Car Wash | Мийка", callback_data=f"{loc_key}_carwash")],
            [InlineKeyboardButton("🔧 Tire Service | Шиномонтаж", callback_data=f"{loc_key}_tire")],
            [InlineKeyboardButton("⬅️ Back | Назад", callback_data="contacts")]
        ]
    elif data.endswith("_carwash"):
        loc_key = data.split("_")[0]
        text = "🧼 Car Washes | Мийки:\n\n"
        for wash in LOCATIONS[loc_key]["car_washes"]:
            text += f"{wash['name']}\nPhone: {wash['phone']}\nMap: {wash['map']}\n\n"
        keyboard = [[InlineKeyboardButton("⬅️ Back | Назад", callback_data=f"contact_{loc_key}")]]
    elif data.endswith("_tire"):
        loc_key = data.split("_")[0]
        text = "🔧 Tire Services | Шиномонтажі:\n\n"
        for tire in LOCATIONS[loc_key]["tire_services"]:
            text += f"{tire['name']}\nPhone: {tire['phone']}\nMap: {tire['map']}\n\n"
        keyboard = [[InlineKeyboardButton("⬅️ Back | Назад", callback_data=f"contact_{loc_key}")]]
    else:
        return

    reply_markup = InlineKeyboardMarkup(keyboard)
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(text=text, reply_markup=reply_markup)

#===================================================================CONTACTS END===================================================











# =================== Main ===================
def main():
    app = Application.builder().token(TOKEN).build()


    ldr_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(ldr_request_type_callback, pattern="^(flat_tire|other_request)$")],
    states={
        SERIAL_LDR: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input_ldr)],
        ODOMETER_LDR: [MessageHandler(filters.TEXT & ~filters.COMMAND, odometer_input_ldr)],
        ALLOCATION_LDR: [CallbackQueryHandler(allocation_input_ldr)],
        TEAM_NUMBER_LDR: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input_ldr)],
        USER_LDR: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input_ldr)],
        DESCRIPTION_LDR: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input_ldr)],
        OTHER_REQUEST_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, ldr_other_request_input)],
    },
    fallbacks=[
        CommandHandler("cancel", cancel),
        CallbackQueryHandler(cancel, pattern="cancel")
    ],
    per_user=True,
    conversation_timeout=300
)




    



    # MFR Conversation
    mfr_conv = ConversationHandler(
    entry_points=[CallbackQueryHandler(mfr_callback, pattern="mfr")],
    states={
        MFR_ALLOCATION: [
            CallbackQueryHandler(mfr_location_selection, pattern="^(Shyroke|Mykolaiv|Kyiv|Sumy/Romny)$"),
            CallbackQueryHandler(allocation_input_mfr)
        ],
        MFR_MODEL_SELECTION: [
            CallbackQueryHandler(model_input_mfr, pattern="^(brand_.*|back_to_brands|.*)$")
        ],
        MFR_SERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input_mfr)],
        MFR_ODOMETER: [MessageHandler(filters.TEXT & ~filters.COMMAND, odometer_input_mfr)],
        MFR_TEAM_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input_mfr)],
        MFR_USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input_mfr)],
        MFR_DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input_mfr)],
    },
    fallbacks=[
        CommandHandler("cancel", cancel),
        CallbackQueryHandler(cancel, pattern="cancel")
    ],
    per_user=True,
    conversation_timeout=300
    )


# ======================== HANDLERS ========================

    # Conversation handlers / form handlers
    app.add_handler(mfr_conv)
    app.add_handler(ldr_conv)
    app.add_handler(inspection_handler)
    # app.add_handler(other_questions_conv)

    # Команды
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("add_user", add_user))
    app.add_handler(CommandHandler("remove_user", remove_user))
    app.add_handler(CommandHandler("list_users", list_users))

    # Главная кнопка (main menu)
    app.add_handler(CallbackQueryHandler(start_button_callback, pattern="^main_menu$"))

    # LDR / MFR
    app.add_handler(CallbackQueryHandler(ldr_callback, pattern="^ldr$"))
    app.add_handler(CallbackQueryHandler(mfr_callback, pattern="^mfr$"))

    # Contacts (главное и локации)
    app.add_handler(CallbackQueryHandler(contacts_callback, pattern="^contacts$"))
    app.add_handler(
        CallbackQueryHandler(
            contact_location_callback,
            pattern="^(contact_shyroke|contact_mykolaiv|shyroke_carwash|shyroke_tire|mykolaiv_carwash|mykolaiv_tire|back)$"
        )
    )

    # Accident procedure (ДТП)
    app.add_handler(CallbackQueryHandler(accident_procedure_callback, pattern="^accident_procedure$"))
    app.add_handler(CallbackQueryHandler(accident_procedure_ua_callback, pattern="^accident_procedure_ua$"))
    app.add_handler(CallbackQueryHandler(accident_procedure_en_callback, pattern="^accident_procedure_en$"))

    # Cancel
    app.add_handler(CallbackQueryHandler(cancel, pattern="^cancel$"))












    app.run_polling()

if __name__ == "__main__":
    main()