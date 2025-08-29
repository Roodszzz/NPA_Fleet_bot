import logging
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ConversationHandler, filters, ContextTypes
)
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from openpyxl.styles import Alignment
from io import BytesIO
from datetime import datetime
from googletrans import Translator
from unidecode import unidecode
import base64
import re
import os
from dotenv import load_dotenv



load_dotenv()  # Загружаем .env
TOKEN = os.getenv("TOKEN")


logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
#==========================================================MFR======================================================================














#==========================================================MFR END======================================================================
def get_workbook():
    # Берем путь относительно текущего скрипта
    current_dir = os.path.dirname(__file__)
    file_path = os.path.join(current_dir, "excel", "LDR.xlsx")  # путь к Excel файлу в папке excel проекта
    return load_workbook(file_path)




# ===== Встроенный логотип =====

def get_logo_bytes():
    # Берем путь относительно текущего скрипта
    current_dir = os.path.dirname(__file__)
    logo_path = os.path.join(current_dir, "logo", "Drive the NPA way.png")
    with open(logo_path, "rb") as f:
        return BytesIO(f.read())


# ===== Состояния =====
SERIAL, ALLOCATION, TEAM_NUMBER, USER, DESCRIPTION = range(5)
translator = Translator()




# Менеджеры по локации
managers = {
    "Shyroke": "manager_shyroke@example.com",
    "Mykolaiv": "manager_mykolaiv@example.com"
}


async def translate_to_en(text: str) -> str:
    result = await translator.translate(text, dest='en')
    return result.text

# ===== Вспомогательные функции =====
def set_cell(ws, cell, value):
    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

def auto_adjust(ws, cells):
    for cell in cells:
        value = ws[cell].value
        if value:
            col_letter = ''.join(filter(str.isalpha, cell))
            ws.column_dimensions[col_letter].width = max(
                ws.column_dimensions[col_letter].width or 10,
                len(str(value)) + 2
            )
            ws.row_dimensions[ws[cell].row].height = max(
                ws.row_dimensions[ws[cell].row].height or 15,
                15
            )


            





# ===== Главное меню =====
async def main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard_inline = [
        [InlineKeyboardButton("LDR / ЛДР", callback_data="ldr")],
        [InlineKeyboardButton("MFR / МФР", callback_data="mfr")],
        [InlineKeyboardButton("VAR / ВАР", callback_data="var")],
        [InlineKeyboardButton("Contacts / Контакти", callback_data="contacts")],
        [InlineKeyboardButton("Other questions / Інші питання", callback_data="other_questions")]
    ]
    reply_markup_inline = InlineKeyboardMarkup(keyboard_inline)
    text = ("Hello! This is the NPA Fleet bot 🚗\nI can help you create reports for vehicles.\n"
            "Привіт! Це бот NPA Fleet 🚗\nЯ допоможу вам створювати звіти по автомобілях.\n"
            "What are you interested in today? / Що вас цікавить сьогодні?")

    if update.callback_query:
        await update.callback_query.answer()
        try:
            await update.callback_query.message.delete()
        except:
            pass
        await update.callback_query.message.reply_text(text=text, reply_markup=reply_markup_inline)
    elif update.message:
        await update.message.reply_text(text=text, reply_markup=reply_markup_inline)

# ===== Start =====
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    logo_bytes = get_logo_bytes()
    logo_file = InputFile(logo_bytes, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start / Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.message:
        await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗\nЛаскаво просимо в NPA Fleet бот", reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗\nЛаскаво просимо в NPA Fleet бот", reply_markup=reply_markup)

async def start_button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await main_menu(update, context)

# ===== Cancel =====


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Очищаем данные пользователя
    context.user_data.clear()

    if update.callback_query:
        # Отвечаем на callback_query, чтобы убрать "часики"
        await update.callback_query.answer()
        # Пытаемся удалить старое сообщение
        try:
            await update.callback_query.message.delete()
        except:
            pass

    # Возврат в главное меню
    await main_menu(update, context)

    return ConversationHandler.END



# ===== LDR =====
async def ldr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    keyboard = [
        [InlineKeyboardButton("Flat tire / Пошкоджене колесо", callback_data="flat_tire")],
        [InlineKeyboardButton("Wipers replacement / Заміна дворників", callback_data="wipers")],
        [InlineKeyboardButton("Other request / Інше звернення", callback_data="other_request")],
        [InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    try: await query.message.delete()
    except: pass
    await query.message.reply_text("Choose request type / Виберіть тип звернення:", reply_markup=reply_markup)




async def ldr_request_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "cancel":
        return await cancel(update, context)
    if data == "other_request":
        keyboard = [[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]
        try: await query.message.delete()
        except: pass
        await query.message.reply_text("You chose: Other request / Ви обрали: Інше звернення. Function in progress / Функція ще в розробці", reply_markup=InlineKeyboardMarkup(keyboard))
        return ConversationHandler.END

    context.user_data['wb'] = get_workbook()
    context.user_data['ws'] = context.user_data['wb'].active
    ws = context.user_data['ws']
    if data == "flat_tire":
        set_cell(ws, "B4", "Flat tyre / Пошкоджене колесо")
    elif data == "wipers":
        set_cell(ws, "B4", "Wipers replacement / Заміна дворників")
    set_cell(ws, "D4", "Serial / ID / Серійний номер / ID")

    keyboard = [
        [InlineKeyboardButton("Shyroke", callback_data="Shyroke")],
        [InlineKeyboardButton("Mykolaiv", callback_data="Mykolaiv")],
        [InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]
    ]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text("Select vehicle location / Оберіть локацію автомобіля:", reply_markup=InlineKeyboardMarkup(keyboard))
    return ALLOCATION

# ===== Ввод данных =====



async def serial_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().upper()  # переводим в верхний регистр
    text = text.replace(" ", "")  # убираем пробелы

    # Если пользователь ввел без дефиса, например AA12, добавим дефис автоматически
    if re.fullmatch(r"[A-Z]{2}\d{2}", text):
        text = text[:2] + "-" + text[2:]

    # проверка формата: две буквы - дефис - две цифры
    if not re.fullmatch(r"[A-Z]{2}-\d{2}", text):
        await update.message.reply_text(
            "❌ Неверный формат номера авто. Формат должен быть: AA-12\nTry again / Спробуйте ще раз:"
        )
        return SERIAL

    ws = context.user_data['ws']
    set_cell(ws, "D4", text)  # записываем нормализованный номер

    keyboard = [
        [InlineKeyboardButton("LOGS", callback_data="LOGS")],
        [InlineKeyboardButton("MTT", callback_data="MTT")],
        [InlineKeyboardButton("MDD", callback_data="MDD")],
        [InlineKeyboardButton("TFM", callback_data="TFM")],
        [InlineKeyboardButton("QA", callback_data="QA")],
        [InlineKeyboardButton("NTS", callback_data="NTS")],
        [InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]
    ]
    await update.message.reply_text(
        "Choose Allocation / Оберіть Allocation:", 
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION



async def allocation_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    selection = query.data
    if selection == "cancel":
        return await cancel(update, context)
    ws = context.user_data['ws']

    if selection in ["Shyroke", "Mykolaiv"]:
        context.user_data['location'] = selection
        set_cell(ws, "B6", selection)
        try: await query.message.delete()
        except: pass
        await query.message.reply_text("Enter vehicle number or call sign (e.g. HP-12) / Введіть номер авто або call sign (напр. HP-12):", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]))
        return SERIAL

    if selection.upper() in ["NTS", "MTT", "MDD"]:
        context.user_data['allocation'] = selection.upper()
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(f"Enter team number for {selection.upper()} / Введіть номер команди для {selection.upper()}:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]))
        return TEAM_NUMBER

    set_cell(ws, "D6", selection)
    try: await query.message.delete()
    except: pass
    await query.message.reply_text("Enter your full name / Введіть ваше ПІБ:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]))
    return USER

async def team_number_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ Team number must be a number / ❌ Номер команди повинен бути числом. Try again / Спробуйте ще раз:")
        return TEAM_NUMBER
    ws = context.user_data['ws']
    allocation = context.user_data['allocation']
    set_cell(ws, "D6", f"{allocation}-{text}")
    await update.message.reply_text("Enter your full name / Введіть ваше ПІБ:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]))
    return USER



async def user_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text(
            "❌ You did not enter your name / ❌ Ви не ввели ПІБ. Try again / Спробуйте ще раз:"
        )
        return USER

    user_name_latin = unidecode(text)
    ws = context.user_data['ws']

    # Записываем данные пользователя
    set_cell(ws, "F4", user_name_latin)
    set_cell(ws, "A10", user_name_latin)

    # Подставляем фамилию менеджера по локации
    managers_fa = {
        "Shyroke": "F.A. Oleksandr Rudnov",
        "Mykolaiv": "F.A. Andriy Padalka"
    }
    location = context.user_data.get('location')
    manager_fa = managers_fa.get(location, "F.A. Unknown")
    set_cell(ws, "D10", manager_fa)

    # Дата
    today_str = datetime.now().strftime("%Y-%m-%d")
    set_cell(ws, "B10", today_str)

    await update.message.reply_text(
        "Briefly describe the situation / Коротко опишіть ситуацію:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
    )
    return DESCRIPTION




#-=======================MFR111+===========================

# ===== MFR модели авто =====
MFR_MODELS = [
    "Toyota Hilux",
    "Toyota Land Cruiser",
    "Ford Transit",
    "Ford Ranger",
    "Mitsubishi L200",
    "Volkswagen",
    "Renault Duster"
]

# ===== MFR старт =====
async def mfr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Клавиатура для выбора модели авто
    keyboard = [[InlineKeyboardButton(model, callback_data=model)] for model in MFR_MODELS]
    keyboard.append([InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")])

    try: await query.message.delete()
    except: pass

    await query.message.reply_text(
        "Select vehicle model / Оберіть модель авто:", 
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return SERIAL  # следующий шаг – ввод серийного номера/ID

# ===== Серийный номер / ID =====
async def mfr_serial_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().upper().replace(" ", "")
    ws = context.user_data['ws']

    # Заглушка записи модели в Excel
    model = context.user_data.get("vehicle")
    set_cell(ws, "A1", model or "MODEL")  # сюда потом вставишь нужную ячейку

    # Проверка формата номера (пример)
    if not re.fullmatch(r"[A-Z]{2}-\d{2}", text):
        await update.message.reply_text(
            "❌ Invalid format. Example: AA-12 / Неправильний формат. Приклад: AA-12"
        )
        return SERIAL

    context.user_data['serial'] = text
    set_cell(ws, "B1", text)  # заглушка для серийного номера
    # Далее – выбор локации
    keyboard = [
        [InlineKeyboardButton("Shyroke", callback_data="Shyroke")],
        [InlineKeyboardButton("Mykolaiv", callback_data="Mykolaiv")],
        [InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]
    ]
    await update.message.reply_text(
        "Select vehicle location / Оберіть локацію автомобіля:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION

# ===== Локация и Allocation =====
async def mfr_allocation_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    selection = query.data
    ws = context.user_data['ws']

    if selection == "cancel":
        return await cancel(update, context)

    context.user_data['location'] = selection
    set_cell(ws, "C1", selection)  # заглушка для локации

    # Если NTS, MTT, MDD – спрашиваем номер команды
    allocation_options = ["NTS", "MTT", "MDD"]
    if selection.upper() in allocation_options:
        context.user_data['allocation'] = selection
        await query.message.reply_text(
            f"Enter team number for {selection} / Введіть номер команди для {selection}:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
        )
        return TEAM_NUMBER
    else:
        # Иначе ввод имени
        await query.message.reply_text(
            "Enter your full name / Введіть ПІБ:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
        )
        return USER

# ===== Номер команды =====
async def mfr_team_number_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ Team number must be a number / ❌ Номер команди повинен бути числом")
        return TEAM_NUMBER
    ws = context.user_data['ws']
    allocation = context.user_data.get('allocation', 'Unknown')
    set_cell(ws, "D1", f"{allocation}-{text}")  # заглушка для номера команды
    await update.message.reply_text(
        "Enter your full name / Введіть ПІБ:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
    )
    return USER

# ===== Имя пользователя =====
async def mfr_user_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ You did not enter your name / ❌ Ви не ввели ПІБ")
        return USER
    ws = context.user_data['ws']
    user_name = unidecode(text)
    context.user_data['user_name'] = user_name
    set_cell(ws, "E1", user_name)  # заглушка для имени
    await update.message.reply_text(
        "Briefly describe the situation / Коротко опишіть ситуацію:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
    )
    return DESCRIPTION

# ===== Описание =====
async def mfr_description_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
        return DESCRIPTION
    ws = context.user_data['ws']
    set_cell(ws, "F1", text)  # заглушка для описания
    await update.message.reply_text("✅ MFR form completed (Excel placeholders used) / MFR форма заповнена")

    # Очистка данных пользователя
    context.user_data.clear()
    return ConversationHandler.END








#============MFR====================
async def description_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text(
            "❌ Describe the situation / ❌ Опишіть ситуацію. Try again / Спробуйте ще раз:"
        )
        return DESCRIPTION

    # Перевод текста на английский
    description_en = await translate_to_en(text)

    ws = context.user_data['ws']

    # ====== Назначение ячеек для MFR ======
    vehicle = context.user_data.get("vehicle", "VEHICLE/MACHINE")        # Пользовательский ввод
    reg_number = ws["D4"].value or "CAR"                                 # Reg / Serial No.
    driver_name = context.user_data.get("user_name", "Unknown")           # Driver / Operator
    location = context.user_data.get("location", "Unknown")               # Локація
    allocation = context.user_data.get("allocation", "Unknown")           # Розподіл

    set_cell(ws, "C5", vehicle)        # Автомобіль або Машина
    set_cell(ws, "F5", reg_number)     # Реєстр. Номер
    set_cell(ws, "I5", driver_name)    # Водій / Оператор
    set_cell(ws, "C8", location)       # Локація
    set_cell(ws, "F8", allocation)     # Розподіл
    set_cell(ws, "B15", description_en) # Опис несправності або необхідне обслуговування
    set_cell(ws, "B22", driver_name)   # дубляж ФИО
    set_cell(ws, "F22", "Manager Name") # Соответствующий менеджер (можно заменить словарём по location)

    # Автоподгон ширины и высоты для всех используемых ячеек
    auto_adjust(ws, ["C5","F5","I5","C8","F8","B15","B22","F22"])

    # ==== Вставляем логотип в Excel ====
    logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
    img = Image(logo_path)
    img.width, img.height = 396, 72
    ws.add_image(img, "A1")

    # ==== Формируем динамическое имя файла ====
    plate = reg_number
    filename = f"MFR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # Сохраняем workbook в поток памяти
    file_stream = BytesIO()
    ws.parent.save(file_stream)
    file_stream.seek(0)

    # Отправка пользователю
    await update.message.reply_document(document=file_stream, filename=filename)
    await update.message.reply_text("✅ MFR File sent / ✅ Файл MFR відправлено")

    # Отправка менеджерам
    location = context.user_data.get('location')
    if location:
        email_to = managers_mfr.get(location)
        if email_to:
            logging.info(f"[TEST MODE] Excel would be sent to {email_to} for location {location}")
            for admin_id in [int(os.getenv("ADMIN_ID"))]:  # заглушка
                file_stream.seek(0)
                await context.bot.send_document(chat_id=admin_id, document=file_stream, filename=filename)

    # Очистка данных пользователя
    context.user_data.clear()

    # Стартовое окно с логотипом для Telegram
    logo_bytes_start = get_logo_bytes()
    logo_file = InputFile(logo_bytes_start, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start / Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_photo(
        photo=logo_file,
        caption="Welcome to NPA Fleet bot 🚗\nЛаскаво просимо в NPA Fleet бот",
        reply_markup=reply_markup
    )

    return ConversationHandler.END






async def description_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text(
            "❌ Describe the situation / ❌ Опишіть ситуацію. Try again / Спробуйте ще раз:"
        )
        return DESCRIPTION

    # Перевод текста на английский
    text_en = await translate_to_en(text)

    ws = context.user_data['ws']
    set_cell(ws, "A9", text_en)
    auto_adjust(ws, ["B4","D4","B6","D6","F4","A10","A9","B10"])

    # ==== Вставляем логотип в Excel ====
    logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
    img = Image(logo_path)
    img.width = 396  # ширина в пикселях
    img.height = 72  # высота в пикселях
    ws.add_image(img, "A1")  # вставляем в ячейку A1

    # ==== Формируем динамическое имя файла ====
    # если у тебя сохраняется номер машины в user_data (например context.user_data['plate']),
    # то можно его вставить

    plate = ws["D4"].value or "CAR"
    filename = f"LDR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    

    # Сохраняем workbook в поток памяти
    file_stream = BytesIO()
    ws.parent.save(file_stream)
    file_stream.seek(0)

    # Отправка пользователю
    await update.message.reply_document(document=file_stream, filename=filename)
    await update.message.reply_text("✅ File sent / ✅ Файл відправлено")

    # Отправка менеджерам по локации
    location = context.user_data.get('location')
    if location:
        ADMIN_ID = int(os.getenv("ADMIN_ID"))
        tg_users = {
            "Shyroke": [ADMIN_ID],
            "Mykolaiv": [ADMIN_ID]  # заглушка
        }  
        
        for user_id in tg_users.get(location, []):
            file_stream.seek(0)
            await context.bot.send_document(chat_id=user_id, document=file_stream, filename=filename)
        
        

    # Очистка данных пользователя
    context.user_data.clear()

    # Стартовое окно с логотипом для Telegram
    logo_bytes_start = get_logo_bytes()
    logo_file = InputFile(logo_bytes_start, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start / Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_photo(
        photo=logo_file,
        caption="Welcome to NPA Fleet bot 🚗\nЛаскаво просимо в NPA Fleet бот",
        reply_markup=reply_markup
    )

    return ConversationHandler.END




# ===== Заглушки =====
async def generic_stub(update: Update, context: ContextTypes.DEFAULT_TYPE, name="Function"):
    query = update.callback_query
    await query.answer()
    keyboard = [[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(f"You selected / Ви обрали {name}. Function in progress / Функція ще в розробці", reply_markup=InlineKeyboardMarkup(keyboard))

async def mfr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    return await generic_stub(update, context, "MFR / МФР")
async def var_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    return await generic_stub(update, context, "VAR / ВАР")
async def contacts_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    return await generic_stub(update, context, "Contacts / Контакти")
async def other_questions_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    return await generic_stub(update, context, "Other questions / Інші питання")

# ===== Запуск =====
def main():
    app = Application.builder().token(TOKEN).build()
#========================================MFR============================================================
    mfr_conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(mfr_request_type_callback, pattern="^(mfr_opt1|mfr_opt2|Shyroke|Mykolaiv)$")],
        states={
            SERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input)],
            ALLOCATION: [CallbackQueryHandler(allocation_input, pattern="^(?!cancel$).*$")],
            TEAM_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input)],
            USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input)],
            DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input_mfr)],  # отдельная функция
        },
        fallbacks=[CommandHandler("cancel", cancel), CallbackQueryHandler(cancel, pattern="cancel")],
        per_user=True,
        conversation_timeout=900
    )
    
#========================================LDR============================================================
    conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(ldr_request_type_callback, pattern="^(flat_tire|wipers|other_request|Shyroke|Mykolaiv)$")],
        states={
            SERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input)],
            ALLOCATION: [CallbackQueryHandler(allocation_input, pattern="^(?!cancel$).*$")],
            TEAM_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input)],
            USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input)],
            DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input)],
        },
        fallbacks=[CommandHandler("cancel", cancel), CallbackQueryHandler(cancel, pattern="cancel")],
        per_user=True,
        conversation_timeout=900
    )
    app.add_handler(mfr_conv_handler)
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(start_button_callback, pattern="main_menu"))
    app.add_handler(CallbackQueryHandler(ldr_callback, pattern="ldr"))
    app.add_handler(conv_handler)
    app.add_handler(CallbackQueryHandler(mfr_callback, pattern="mfr"))
    app.add_handler(CallbackQueryHandler(var_callback, pattern="var"))
    app.add_handler(CallbackQueryHandler(contacts_callback, pattern="contacts"))
    app.add_handler(CallbackQueryHandler(other_questions_callback, pattern="other_questions"))
    app.add_handler(CallbackQueryHandler(cancel, pattern="cancel"))

    app.run_polling()

if __name__ == "__main__":
    main()
