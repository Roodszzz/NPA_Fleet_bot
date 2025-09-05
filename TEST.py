import logging
import os
import re
from io import BytesIO
from datetime import datetime
from unidecode import unidecode
from dotenv import load_dotenv
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ConversationHandler, filters, ContextTypes
)

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from googletrans import Translator
import math




# =================== Загрузка переменных ===================
load_dotenv()
TOKEN = os.getenv("TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

# =================== Логирование ===================
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# =================== Состояния ===================
SERIAL, ALLOCATION, TEAM_NUMBER, USER, DESCRIPTION = range(5)
translator = Translator()



# =================== Вспомогательные функции ===================
def get_workbook(report_type="LDR"):
    current_dir = os.path.dirname(__file__)
    if report_type.upper() == "MFR":
        filename = "MFR.xlsx"
    else:
        filename = "LDR.xlsx"
    return load_workbook(os.path.join(current_dir, "excel", filename))


def get_logo_bytes():
    current_dir = os.path.dirname(__file__)
    with open(os.path.join(current_dir, "logo", "Drive the NPA way.png"), "rb") as f:
        return BytesIO(f.read())

async def translate_to_en(text: str) -> str:
    result = await translator.translate(text, dest='en')
    return result.text

# def set_cell(ws, cell, value):
#     ws[cell] = value
#     ws[cell].alignment = Alignment(horizontal="center", vertical="center")

# def auto_adjust(ws, cells):
#     for cell in cells:
#         value = ws[cell].value
#         if value:
#             col_letter = ''.join(filter(str.isalpha, cell))
#             ws.column_dimensions[col_letter].width = max(
#                 ws.column_dimensions[col_letter].width or 10,
#                 len(str(value)) + 2
#             )
#             ws.row_dimensions[ws[cell].row].height = max(
#                 ws.row_dimensions[ws[cell].row].height or 15,
#                 15
#             )


def set_cell(ws, cell, value):
    try:
        ws[cell].value = value
    except AttributeError:
        # если попали в объединённую ячейку, ищем верхнюю левую
        for merged_range in ws.merged_cells.ranges:
            if cell in merged_range:
                top_left = merged_range.min_row, merged_range.min_col
                ws.cell(row=top_left[0], column=top_left[1], value=value)
                break
def auto_adjust(ws, cells):
    for cell in cells:
        value = ws[cell].value
        if value:
            col_letter = ''.join(filter(str.isalpha, cell))
            ws.column_dimensions[col_letter].width = max(
                ws.column_dimensions[col_letter].width or 10,
                len(str(value)) + 2
            )
            ws.row_dimensions[ws[cell].row].height = max(
                ws.row_dimensions[ws[cell].row].height or 15,
                15
            )



# =================== Главное меню ===================
async def main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("LDR (Lost / Damage) | Втрачено або пошкоджено", callback_data="ldr")],
        [InlineKeyboardButton("MFR (Mechanical failure) | Механічне пошкодження авто", callback_data="mfr")],
        [InlineKeyboardButton("Contacts | Контакти", callback_data="contacts")],
        #[InlineKeyboardButton("Other questions | Інші питання", callback_data="other_questions")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    text = ("🇬🇧 EN\n"
            "Hello! This is the NPA Fleet bot 🚗\n"
            "I can help you create reports for vehicles.\n\n"
            "🇺🇦 UA\n"
            "Привіт! Це бот NPA Fleet 🚗\n"
            "Я допоможу вам створювати звіти по автомобілях.\n\n"
            "What are you interested in today? / Що вас цікавить сьогодні?"
            )
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.message.delete()
        except: pass
        await update.callback_query.message.reply_text(text=text, reply_markup=reply_markup)
    else:
        await update.message.reply_text(text=text, reply_markup=reply_markup)





# если нужно ограничить доступ — ставишь True
RESTRICTED_MODE = False






# =================== Старт ===================

ALLOWED_USERS = {
    5077758580: "Oleksandr Rudnov",
    6093640376: "Roman Kucherevskyi",
    787549014: "Anastasia Vesloguzova",
    513781701: "Dmytro Safonenko",
    528557238: "Vladyslav Prikhodko",
    702797267: "Maksym Shevchenko",
}


def restricted(func):
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE, *args, **kwargs):
        user_id = update.effective_user.id
        if user_id not in ALLOWED_USERS:
            if update.message:
                await update.message.reply_text("⛔ Доступ заборонений")
            elif update.callback_query:
                await update.callback_query.answer()
                await update.callback_query.message.reply_text("⛔ Доступ заборонений")
            return
        return await func(update, context, *args, **kwargs)
    return wrapper

#@restricted
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    if RESTRICTED_MODE:
        if user_id not in ALLOWED_USERS:
            if update.message:
                await update.message.reply_text("⛔ Доступ заборонений")
            elif update.callback_query:
                await update.callback_query.answer()
                await update.callback_query.message.reply_text("⛔ Доступ заборонений")
            return  # прекращаем выполнение
        


#@restricted
# async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
#     user_id = update.effective_user.id
#     if user_id not in ALLOWED_USERS:
#         if update.message:
#             await update.message.reply_text("⛔ Доступ заборонений")
#         elif update.callback_query:
#             await update.callback_query.answer()
#             await update.callback_query.message.reply_text("⛔ Доступ заборонений")
#         return  # прекращаем выполнение, дальше ничего не делаем

    # очищаем user_data
    context.user_data.clear()

    # готовим фото
    logo_bytes = get_logo_bytes()
    logo_file = InputFile(logo_bytes, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)

    # отправляем фото
    if update.message:
        await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

#@restricted
async def start_button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await main_menu(update, context)

# =================== Cancel ===================
#@restricted
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.message.delete()
        except: pass
    await main_menu(update, context)
    return ConversationHandler.END


MANAGERS = {
    "Shyroke": [ADMIN_ID],
    "Mykolaiv": [6093640376],
}













# ================================================================== LDR ========================================================================================================
SERIAL = 1
ALLOCATION = 2
TEAM_NUMBER = 3
USER = 4
DESCRIPTION = 5
OTHER_REQUEST_INPUT = 6

#@restricted
async def ldr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    keyboard = [
        [InlineKeyboardButton("Flat tire | Пошкоджене колесо", callback_data="flat_tire")],
        #[InlineKeyboardButton("Wipers replacement | Заміна дворників", callback_data="wipers")],
        #[InlineKeyboardButton("Driver's card | Водійська карта", callback_data="Drivers_card")],
        [InlineKeyboardButton("Other damage | Інше пошкодження", callback_data="other_request")],
        [InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]
    ]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text("Choose request type | Виберіть тип звернення:", reply_markup=InlineKeyboardMarkup(keyboard))


#@restricted
from telegram import InlineKeyboardButton, InlineKeyboardMarkup

async def ldr_request_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "cancel":
        return await cancel(update, context)

    context.user_data['wb'] = get_workbook("LDR")
    context.user_data['ws'] = context.user_data['wb'].active
    ws = context.user_data['ws']

    if data == "other_request":
        try:
            await query.message.delete()
        except:
            pass

        # создаём кнопку Cancel
        cancel_button = InlineKeyboardMarkup([
            [InlineKeyboardButton("Cancel ❌", callback_data="cancel")]
        ])

        await query.message.reply_text(
            "Please indicate what is damaged: \nВкажіть, що пошкоджено:",
            reply_markup=cancel_button
        )
        
        return OTHER_REQUEST_INPUT  # <-- возвращаем константу, а не строку


    # Старые варианты кнопок с готовыми фразами
    if data == "flat_tire":
        set_cell(ws, "C7", "Flat tyre")
    elif data == "wipers":
        set_cell(ws, "C7", "Wipers replacement")
    elif data == "Drivers_card":
        set_cell(ws, "C7", "Driver's card")    

    set_cell(ws, "F7", "Serial / ID / Серійний номер / ID")

    keyboard = [
        [InlineKeyboardButton("Shyroke", callback_data="Shyroke")],
        [InlineKeyboardButton("Mykolaiv", callback_data="Mykolaiv")],
        [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
    ]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(
        "Select vehicle location | Оберіть локацію автомобіля:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION



# Новый хэндлер для ввода текста пользователем
from googletrans import Translator

translator = Translator()

async def translate_to_en(text: str) -> str:
    # Асинхронно вызываем перевод
    translated = await translator.translate(text, dest='en')
    return translated.text

#@restricted
async def ldr_other_request_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_text = update.message.text.strip()
    if not user_text:
        await update.message.reply_text("❌ Please type your request / ❌ Введіть ваше звернення")
        return OTHER_REQUEST_INPUT

    ws = context.user_data['ws']

    # Перевод на английский
    translated_text = await translate_to_en(user_text)

    # Записываем перевод в Excel
    set_cell(ws, "C7", translated_text)
    set_cell(ws, "F7", "Serial / ID / Серійний номер / ID")

    keyboard = [
        [InlineKeyboardButton("Shyroke", callback_data="Shyroke")],
        [InlineKeyboardButton("Mykolaiv", callback_data="Mykolaiv")],
        [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
    ]
    await update.message.reply_text(
        "Select vehicle location | Оберіть локацію автомобіля:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION

# =================== Ввод данных ===================











async def serial_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().upper()  # переводим в верхний регистр
    text = text.replace(" ", "")  # убираем пробелы

    # Если пользователь ввел без дефиса, например AA12, добавим дефис автоматически
    if re.fullmatch(r"[A-Z]{2}\d{2}", text):
        text = text[:2] + "-" + text[2:]

    # проверка формата: две буквы - дефис - две цифры
    if not re.fullmatch(r"[A-Z]{2}-\d{2}", text):
        await update.message.reply_text(
            "❌ Невірный формат номера авто. Формат повинен бути:(напр. HP-12)\nTry again / Спробуйте ще раз:"
        )
        return SERIAL

    ws = context.user_data['ws']
    set_cell(ws, "F7", text)

    # Первый уровень Allocation
    keyboard = [
        [InlineKeyboardButton(x, callback_data=x)] for x in ["MTT","MDD","MECH","NTS","OPS/SUPP"]
    ]
    keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])

    await update.message.reply_text(
        "Choose Allocation | Оберіть Розподіл:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION


async def allocation_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    selection = query.data

    # Проверяем наличие workbook
    if 'ws' not in context.user_data:
        await query.message.reply_text(
            "❌ Please start the request from the beginning using /start\n❌ Будь ласка, почніть звернення заново за допомогою /start"
        )
        return ConversationHandler.END

    ws = context.user_data['ws']

    # Обработка отмены
    if selection == "cancel":
        return await cancel(update, context)

    # Локации Shyroke / Mykolaiv
    if selection in ["Shyroke", "Mykolaiv"]:
        context.user_data['location'] = selection
        set_cell(ws, "C10", selection)
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter vehicle call sign (e.g. HP-12): \nВведіть внутрішній номер авто (напр. HP-12):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return SERIAL

    # Если пользователь выбрал OPS/SUPP — показываем второй уровень кнопок
    if selection == "OPS/SUPP":
        keyboard = [[InlineKeyboardButton(x, callback_data=f"OPS/{x}")] for x in ["STFM","TFM","SUPV","LOGS","IMM","QA"]]
        keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Choose sub-allocation for OPS/SUPP | Оберіть підрозподіл для OPS/SUPP:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return ALLOCATION  # остаёмся на этом же шаге, ждём второй выбор

    # Обработка выбора подкнопки OPS/SUPP
    if selection.startswith("OPS/"):
        allocation_choice = selection.split("/")[1]
        set_cell(ws, "F10", f"{allocation_choice}")
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter your full name | Введіть ваше Ім'я та прізвище:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return USER

    # Если MTT, MDD, NTS — спрашиваем номер команды
    if selection.upper() in ["MTT", "MDD", "NTS"]:
        context.user_data['allocation'] = selection.upper()
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            f"Enter team number for {selection.upper()} | Введіть номер команди для {selection.upper()}:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return TEAM_NUMBER

    # Если MECH — просто записываем в Excel и спрашиваем имя
    if selection.upper() == "MECH":
        set_cell(ws, "D6", "MECH")
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter your full name | Введіть ваше Ім'я та прізвище:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return USER


async def team_number_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ Team number must be a number")
        return TEAM_NUMBER
    ws = context.user_data['ws']
    allocation = context.user_data.get('allocation')
    set_cell(ws, "F10", f"{allocation}-{text}")
    await update.message.reply_text(
        "Enter your full name | Введіть Ім'я та прізвище:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
    )
    return USER


async def user_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ You did not enter your name")
        return USER
    user_name_latin = unidecode(text)
    ws = context.user_data['ws']
    set_cell(ws, "I7", user_name_latin)
    set_cell(ws, "B21", user_name_latin)
    location = context.user_data.get('location')
    manager_fa = {"Shyroke":"F.A. Oleksandr Rudnov","Mykolaiv":"F.A. Andriy Padalka"}.get(location,"F.A. Unknown")
    set_cell(ws, "F21", manager_fa)
    set_cell(ws, "C21", datetime.now().strftime("%Y-%m-%d"))
    await update.message.reply_text(
        "Briefly describe the situation | Коротко опишіть ситуацію:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
    )
    return DESCRIPTION



# =================== Описание ===================




def auto_height_for_cell(ws, cell_address):
    cell = ws[cell_address]
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Получаем ширину колонки в символах (приближённо)
    col_letter = ''.join(filter(str.isalpha, cell_address))
    col_width = ws.column_dimensions[col_letter].width or 10  # если не задано, ставим 10

    # Оценка количества строк: длина текста / ширина колонки
    text_length = len(str(cell.value))
    lines_needed = math.ceil(text_length / col_width)

    # Стандартная высота одной строки ~15
    ws.row_dimensions[cell.row].height = lines_needed * 15

# Пример использования в твоей функции:



async def description_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
        return DESCRIPTION

    text_en = await translate_to_en(text)
    ws = context.user_data['ws']









    # # Вставка текста в A9 с переносом и выравниванием по левому краю
    # cell = ws["A9"]
    # cell.value = text_en
    # from openpyxl.styles import Alignment
    # cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # # Авто-высота строки для A9 с минимальной высотой
    # auto_height_for_cell(ws, "A9", min_height=200)



def split_text(text, words_per_line=12):
    """Разбивает текст на строки примерно по 20 слов"""
    words = text.split()
    return [" ".join(words[i:i+words_per_line]) for i in range(0, len(words), words_per_line)]

async def description_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
        return DESCRIPTION

    text_en = await translate_to_en(text)
    ws = context.user_data['ws']

    # Разбиваем текст на куски
    lines = split_text(text_en, words_per_line=25)

    # вставка текста по строкам
    start_row = 16  # теперь B16
    for i, line in enumerate(lines, start=start_row):
        if i > 20:
            break
        cell = ws[f"B{i}"]
        cell.value = line
        cell.alignment = Alignment(horizontal="left", vertical="top")


    # Подгоняем размеры остальных ячеек
    auto_adjust(ws, ["C7","F7","C10","F10","I7","B21","C21","F21"])


    

    # Логотип
    # logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
    # img = Image(logo_path)
    # img.width, img.height = 1069, 194
    # ws.add_image(img, "A1")

    plate = ws["F7"].value or "CAR"
    filename = f"LDR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # Отправка менеджерам по локации


    # Отправка менеджерам по локации
    location = context.user_data.get("location")
    manager_ids = MANAGERS.get(location, [])
    user_id = update.effective_user.id
    user_name = ALLOWED_USERS.get(user_id, "Unknown")  # получаем имя из словаря

    for manager_id in manager_ids:
        file_stream = BytesIO()
        ws.parent.save(file_stream)
        file_stream.seek(0)
        await context.bot.send_document(chat_id=manager_id, document=file_stream, filename=filename)
        await context.bot.send_message(
            chat_id=manager_id,
            text=f"📄 Новий LDR звіт по локації {location} від {user_name}"
        )

    context.user_data.clear()

    # Уведомление пользователю
    await update.message.reply_text("✅ Your report has been sent! / ✅ Звіт надіслано!")

    # Приветственное фото с кнопкой
    logo_bytes_start = get_logo_bytes()
    logo_file = InputFile(logo_bytes_start, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

    return ConversationHandler.END

# Функция для авто-подгонки высоты строки A9 с минимальной защитой
def auto_height_for_cell(ws, cell_address, min_height=45):
    """Автоматическая высота строки под содержимое, но не меньше min_height"""
    cell = ws[cell_address]
    row = cell.row
    lines = str(cell.value).count('\n') + 1
    # Расчет высоты: 15 пикселей на строку
    height = max(lines * 15, min_height)
    ws.row_dimensions[row].height = height






# =================== Заглушки ===================
async def generic_stub(update: Update, context: ContextTypes.DEFAULT_TYPE, name="Function"):
    query = update.callback_query
    await query.answer()
    keyboard = [[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(f"You selected {name}. Function in progress.", reply_markup=InlineKeyboardMarkup(keyboard))


#=====================================================LDR END=============================================================================














#=====================================================MFR=================================================================================

ALLOCATION, MODEL_SELECTION, SERIAL, TEAM_NUMBER, USER, DESCRIPTION = range(6)

# Начало MFR запроса — спрашиваем локацию
#@restricted
async def mfr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Создаем workbook для MFR
    context.user_data['wb'] = get_workbook("MFR")
    context.user_data['ws'] = context.user_data['wb'].active
    ws = context.user_data['ws']

    # Тип запроса в Excel
    set_cell(ws, "F6", "Serial / ID / Серійний номер / ID")

    # ------------------- Кнопки локации -------------------
    keyboard = [
        [InlineKeyboardButton("Shyroke", callback_data="Shyroke")],
        [InlineKeyboardButton("Mykolaiv", callback_data="Mykolaiv")],
        [InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]
    ]
    # -------------------------------------------------------

    try: 
        await query.message.delete()
    except: 
        pass

    await query.message.reply_text(
        "Select vehicle location | Оберіть локацію автомобіля:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION




# Выбор локации
# ---------- Первый уровень: бренды ----------
async def mfr_location_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    location = query.data

    if location == "cancel":
        return await cancel(update, context)

    ws = context.user_data['ws']
    set_cell(ws, "C9", location)
    context.user_data['location'] = location

    try:
        await query.message.delete()
    except:
        pass

    # бренды
    keyboard = [
        [InlineKeyboardButton("TOYOTA", callback_data="brand_TOYOTA")],
        [InlineKeyboardButton("FORD", callback_data="brand_FORD")],
        [InlineKeyboardButton("MITSUBISHI", callback_data="brand_MITSUBISHI")],
        [InlineKeyboardButton("VOLKSWAGEN", callback_data="brand_VOLKSWAGEN")],
        [InlineKeyboardButton("RENAULT DUSTER", callback_data="RENAULT DUSTER")],
        [InlineKeyboardButton("SKODA KODIAQ", callback_data="SKODA KODIAQ")],
        [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
    ]

    await query.message.reply_text(
        "Select car brand | Оберіть марку авто:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return MODEL_SELECTION




# Выбор модели авто
# ---------- Второй уровень: модели ----------
async def model_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    choice = query.data

    ws = context.user_data['ws']

    # Отмена
    if choice == "cancel":
        return await cancel(update, context)

    # Если нажали бренд -> показать подмодели
    if choice.startswith("brand_"):
        brand = choice.replace("brand_", "")

        if brand == "TOYOTA":
            models = ["Toyota Hilux", "Toyota Land Cruiser"]
        elif brand == "FORD":
            models = ["Ford Ranger", "Ford Transit", "Ford Truck"]
        elif brand == "MITSUBISHI":
            models = ["Mitsubishi L200", "Mitsubishi ASX", "Mitsubishi Outlander"]
        elif brand == "VOLKSWAGEN":
            models = ["Volkswagen T6", "Volkswagen ID.4"]
        else:
            models = []

        keyboard = [[InlineKeyboardButton(m, callback_data=m)] for m in models]
        keyboard.append([InlineKeyboardButton("⬅️ Back | Назад", callback_data="back_to_brands")])
        keyboard.append([InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")])

        try:
            await query.message.delete()
        except:
            pass

        await query.message.reply_text(
            f"Select model of {brand}:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        return MODEL_SELECTION

    # Вернуться на список брендов
    if choice == "back_to_brands":
        return await mfr_location_selection(update, context)

    # Если выбрана конкретная модель
    model_name = choice
    set_cell(ws, "C6", model_name)

    try:
        await query.message.delete()
    except:
        pass

    await query.message.reply_text(
        "Enter vehicle call sign (e.g. HP-12): \nВведіть внутрішній номер авто (напр. HP-12):",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
    )
    return SERIAL




async def serial_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().upper()
    text = text.replace(" ", "")

    if re.fullmatch(r"[A-Z]{2}\d{2}", text):
        text = text[:2] + "-" + text[2:]

    if not re.fullmatch(r"[A-Z]{2}-\d{2}", text):
        await update.message.reply_text(
            "❌ Невірный формат номера авто. Формат повинен бути:(напр. HP-12)\nTry again | Спробуйте ще раз:"
        )
        return SERIAL

    ws = context.user_data['ws']
    set_cell(ws, "F6", text)

    # Первый уровень Allocation
    keyboard = [[InlineKeyboardButton(x, callback_data=x)] for x in ["MTT", "MDD", "MECH", "NTS", "OPS/SUPP"]]
    keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])

    await update.message.reply_text(
        "Choose Allocation | Оберіть Allocation:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION


async def allocation_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if 'ws' not in context.user_data:
        await query.message.reply_text("❌ Please start from /start")
        return ConversationHandler.END

    ws = context.user_data['ws']
    selection = query.data

    # Обработка отмены
    if selection == "cancel":
        return await cancel(update, context)

    # Локации Shyroke / Mykolaiv
    if selection in ["Shyroke", "Mykolaiv"]:
        context.user_data['location'] = selection
        set_cell(ws, "C9", selection)
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter vehicle call sign (e.g. HP-12) | Введіть внутрішній номер авто (напр. HP-12):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return SERIAL

    # Если пользователь выбрал OPS/SUPP — показываем второй уровень кнопок
    if selection == "OPS/SUPP":
        keyboard = [[InlineKeyboardButton(x, callback_data=f"OPS/{x}")] for x in ["STFM","TFM","SUPV","LOGS","IMM","QA"]]
        keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])
        try:
            await query.message.edit_text(
                "Choose sub-allocation for OPS/SUPP | Оберіть підрозподіл для OPS/SUPP:",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        except:
            await query.message.reply_text(
                "Choose sub-allocation for OPS/SUPP | Оберіть підрозподіл для OPS/SUPP:",
                reply_markup=InlineKeyboardMarkup(keyboard)
            )
        return ALLOCATION

    # Обработка выбора подкнопки OPS/SUPP
    if selection.startswith("OPS/"):
        allocation_choice = selection.split("/")[1]
        set_cell(ws, "F9", f"{allocation_choice}")
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter your full name | Введіть ваше ім'я та прізвище:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return USER

    # MTT, MDD, NTS — спрашиваем номер команды
    if selection.upper() in ["MTT", "MDD", "NTS"]:
        context.user_data['allocation'] = selection.upper()
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            f"Enter team number for {selection.upper()} | Введіть номер команди для {selection.upper()}:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return TEAM_NUMBER

    # MECH — просто записываем и спрашиваем имя
    if selection.upper() == "MECH":
        set_cell(ws, "F9", "MECH")
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter your full name | Введіть ваше ім'я та прізвище:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return USER



    # Если MTT, MDD, NTS — спрашиваем номер команды
    if selection.upper() in ["MTT", "MDD", "NTS"]:
        context.user_data['allocation'] = selection.upper()
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            f"Enter team number for {selection.upper()} | Введіть номер команди для {selection.upper()}:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return TEAM_NUMBER

    # Если MECH — просто записываем в Excel и спрашиваем имя
    if selection.upper() == "MECH":
        set_cell(ws, "F9", "MECH")
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter your full name | Введіть ваше ім'я та прізвище:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return USER



async def team_number_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ Team number must be a number")
        return TEAM_NUMBER
    ws = context.user_data['ws']
    allocation = context.user_data.get('allocation')
    set_cell(ws, "F9", f"{allocation}-{text}")
    await update.message.reply_text(
        "Enter your full name | Введіть Ім'я та прізвище:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
    )
    return USER


async def user_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ You did not enter your name")
        return USER
    user_name_latin = unidecode(text)
    ws = context.user_data['ws']
    set_cell(ws, "I6", user_name_latin)
    set_cell(ws, "B23", user_name_latin)
    location = context.user_data.get('location')
    manager_fa = {"Shyroke":"F.A. Oleksandr Rudnov","Mykolaiv":"F.A. Andriy Padalka"}.get(location,"F.A. Unknown")
    set_cell(ws, "F23", manager_fa)
    set_cell(ws, "C23", datetime.now().strftime("%Y-%m-%d"))
    set_cell(ws, "F12", datetime.now().strftime("%Y-%m-%d"))
    await update.message.reply_text(
        "Briefly describe the situation | Коротко опишіть ситуацію:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]])
    )
    return DESCRIPTION



# =================== Описание ===================



async def description_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
        return DESCRIPTION

    text_en = await translate_to_en(text)
    ws = context.user_data['ws']

    # --- Записываем текст в одну ячейку и выравниваем ---
    ws["B16"] = text_en
    ws["B16"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # --- Автоподгонка высоты ячейки под текст ---
    auto_height_for_cell(ws, "B16", min_height=50)

    # --- Подгонка остальных ячеек ---
    auto_adjust(ws, ["F6", "C6", "C9", "F9", "I6", "F23", "C23"])

    # --- Лого ---
    logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
    img = Image(logo_path)
    img.width, img.height = 396, 72
    ws.add_image(img, "B2")

    plate = ws["F6"].value or "CAR"
    filename = f"MFR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # --- Отправка менеджеру ---

    # Отправка менеджерам по локации
    location = context.user_data.get("location")
    manager_ids = MANAGERS.get(location, [])
    user_id = update.effective_user.id
    user_name = ALLOWED_USERS.get(user_id, "Unknown")  # получаем имя из словаря

    for manager_id in manager_ids:
        file_stream = BytesIO()
        ws.parent.save(file_stream)
        file_stream.seek(0)
        await context.bot.send_document(chat_id=manager_id, document=file_stream, filename=filename)
        await context.bot.send_message(
            chat_id=manager_id,
            text=f"📄 Новий MFR звіт по локації {location} від {user_name}"
    )

 

    context.user_data.clear()
    await update.message.reply_text("✅ Your report has been sent! / ✅ Звіт надіслано!")

    # --- Приветственное фото ---
    logo_bytes_start = get_logo_bytes()
    logo_file = InputFile(logo_bytes_start, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

    return ConversationHandler.END



#=============================================================MFR END=============================================================














#===================================================================CONTACTS====================================================

#@restricted
async def contacts_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Если нажата кнопка "Назад", возвращаемся в главное меню
    if query.data == "back":
        try:
            await query.message.delete()
        except:
            pass
        await main_menu(update, context)  # вызываем функцию главного меню
        return

    text = (
        "📌 Locations / Локації:\n"
        "Select a location to see contacts | Оберіть локацію для контактів:"
    )

    keyboard = [
        [
            InlineKeyboardButton("Shyroke | Широке", callback_data="contact_shyroke"),
            InlineKeyboardButton("Mykolaiv | Миколаїв", callback_data="contact_mykolaiv"),
        ],
        [InlineKeyboardButton("❌ Back | Назад", callback_data="back")]
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)
    try:
        await query.message.delete()
    except:
        pass
    await query.message.reply_text(text=text, reply_markup=reply_markup)



#Обработчик конкретной локации
#@restricted
async def contact_location_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "back":
        try: await query.message.delete()
        except: pass
        await main_menu(update, context)
        return

    if data == "contact_shyroke":
        text = (
            "📌 Shyroke | Широке\n"
            "👤 F.A. Oleksandr Rudnov | F.A. Олександр Руднов\n"
            "📞 Phone: +380 431 019 082\n"
            "🌐 Map: https://goo.gl/maps/example1"
        )
        keyboard = [
            [InlineKeyboardButton("Car Wash | Мийка", url="https://goo.gl/maps/carwash_shyroke")],
            [InlineKeyboardButton("Tire Service | Шиномонтаж", url="https://goo.gl/maps/tire_shyroke")],
            [InlineKeyboardButton("❌ Back | Назад", callback_data="contacts")]
        ]
    elif data == "contact_mykolaiv":
        text = (
            "📌 Mykolaiv | Миколаїв\n"
            "👤 F.A. Andriy Padalka | F.A. Андрій Падалка\n"
            "📞 Phone: +380 431 019 083\n"
            "🌐 Map: https://goo.gl/maps/example2"
        )
        keyboard = [
            [InlineKeyboardButton("Car Wash | Мийка", url="https://goo.gl/maps/carwash_mykolaiv")],
            [InlineKeyboardButton("Tire Service | Шиномонтаж", url="https://goo.gl/maps/tire_mykolaiv")],
            [InlineKeyboardButton("❌ Back | Назад", callback_data="contacts")]
        ]

    reply_markup = InlineKeyboardMarkup(keyboard)
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(text=text, reply_markup=reply_markup)


#Контакты по локациям
LOCATIONS = {
    "shyroke": {
        "manager": {
            "name": "Oleksandr Rudnov | Олександр Руднов",
            "phone": "+380987938674",
            "email": "OleRud441@npaid.org"
        },
        "senior_officer": {
            "position": "Senior Fleet Officer",
            "name": "Roman Kucherevskyi",
            "phone": "+380661930132",
            "email": "RomKuc884@npaid.org"
        },
        "car_washes": [
            {"name": "Avtoynhulstroy", "phone": "+380 67 633 1025", "map": "https://www.google.com/maps?cid=3778105884522161440"},
            {"name": "Nova Liniya", "phone": "+380 97 577 2770", "map": "https://www.google.com/maps?cid=1167848751790635382"},
        ],
        "tire_services": [
            {"name": "Avtoynhulstroy", "phone": "+380 67 633 1025", "map": "https://www.google.com/maps?cid=3778105884522161440"},
            {"name": "Nova Liniya", "phone": "+380 97 577 2770", "map": "https://www.google.com/maps?cid=1167848751790635382"},
            {"name": "SHYROKE - Tyre service", "phone": "+380 98 455 8113", "map": "https://maps.app.goo.gl/otgcPE4GaHowdxEj8"},
        ],
    },
    "mykolaiv": {
        "manager": {
            "name": "Andriy Padalka | Андрій Падалка",
            "phone": "+380506008345",
            "email": "AndPad212@npaid.org"
        },
        "senior_officer": {
            "position": "Senior Fleet Officer",
            "name": "Roman Kucherevskyi",
            "phone": "+380661930132",
            "email": "RomKuc884@npaid.org"
        },
        "car_washes": [
            {"name": "Car Wash 1", "phone": "+380 432 000 001", "map": "https://goo.gl/maps/carwash1_mykolaiv"},
            {"name": "Car Wash 2", "phone": "+380 432 000 002", "map": "https://goo.gl/maps/carwash2_mykolaiv"},
        ],
        "tire_services": [
            {"name": "Tire Service 1", "phone": "+380 432 111 001", "map": "https://goo.gl/maps/tire1_mykolaiv"},
            {"name": "Tire Service 2", "phone": "+380 432 111 002", "map": "https://goo.gl/maps/tire2_mykolaiv"},
        ],
    }
}

#@restricted
async def contact_location_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "back":
        try: await query.message.delete()
        except: pass
        await contacts_callback(update, context)
        return

    loc_key = None
    action = None

    if data.startswith("contact_"):
        loc_key = data.split("_")[1]  # shyroke или mykolaiv
        loc_data = LOCATIONS[loc_key]
        manager = loc_data["manager"]
        senior = loc_data["senior_officer"]

        text = (
            f"📌 {loc_key.capitalize()}\n\n"
            f"👤 Fleet Assistant: {manager['name']}\n"
            f"📞 Phone: {manager['phone']}\n"
            f"✉️ Email: {manager['email']}\n\n"
            f"👔 {senior['position']}: {senior['name']}\n"
            f"📞 Phone: {senior['phone']}\n"
            f"✉️ Email: {senior['email']}\n\n"  
        )

        keyboard = [
            [InlineKeyboardButton("🧼 Car Wash | Мийка", callback_data=f"{loc_key}_carwash")],
            [InlineKeyboardButton("🔧 Tire Service | Шиномонтаж", callback_data=f"{loc_key}_tire")],
            [InlineKeyboardButton("❌ Back | Назад", callback_data="contacts")]
        ]
    elif data.endswith("_carwash"):
        loc_key = data.split("_")[0]
        text = "🧼 Car Washes | Мийки:\n\n"
        for wash in LOCATIONS[loc_key]["car_washes"]:
            text += f"{wash['name']}\nPhone: {wash['phone']}\nMap: {wash['map']}\n\n"
        keyboard = [[InlineKeyboardButton("❌ Back | Назад", callback_data=f"contact_{loc_key}")]]
    elif data.endswith("_tire"):
        loc_key = data.split("_")[0]
        text = "🔧 Tire Services | Шиномонтажі:\n\n"
        for tire in LOCATIONS[loc_key]["tire_services"]:
            text += f"{tire['name']}\nPhone: {tire['phone']}\nMap: {tire['map']}\n\n"
        keyboard = [[InlineKeyboardButton("❌ Back | Назад", callback_data=f"contact_{loc_key}")]]
    else:
        return

    reply_markup = InlineKeyboardMarkup(keyboard)
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(text=text, reply_markup=reply_markup)

#===================================================================CONTACTS END===================================================











# =================== Main ===================
def main():
    app = Application.builder().token(TOKEN).build()

# LDR Conversation


#     ldr_conv = ConversationHandler(
#     entry_points=[CallbackQueryHandler(ldr_request_type_callback, pattern="^(flat_tire|wipers|Drivers_card|other_request)$")],
#     states={
#         SERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input_ldr)],
#         ALLOCATION: [CallbackQueryHandler(allocation_input_ldr)],
#         TEAM_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input_ldr)],
#         USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input_ldr)],
#         DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input_ldr)],
#         OTHER_REQUEST_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, ldr_other_request_input)],  # новый шаг
#     },
#     fallbacks=[CommandHandler("cancel", cancel), CallbackQueryHandler(cancel, pattern="cancel")],
#     per_user=True
# )



# # MFR Conversation
    
#     mfr_conv = ConversationHandler(
#     entry_points=[CallbackQueryHandler(mfr_callback, pattern="mfr")],
#     states={
#         # ------------------- Локации -------------------
#         ALLOCATION: [
#             CallbackQueryHandler(mfr_location_selection, pattern="^(Shyroke|Mykolaiv)$"),
#             CallbackQueryHandler(allocation_input_mfr)  # все остальные аллокации
#         ],

#         # ------------------- Бренды и модели -------------------
#         MODEL_SELECTION: [
#             CallbackQueryHandler(model_input_mfr, pattern="^(brand_.*|back_to_brands|.*)$")
#         ],

#         # ------------------- Ввод данных -------------------
#         SERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input_mfr)],
#         TEAM_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input_mfr)],
#         USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input_mfr)],
#         DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input_mfr)],
#     },
#     fallbacks=[
#         CommandHandler("cancel", cancel),
#         CallbackQueryHandler(cancel, pattern="cancel")
#     ],
#     per_user=True
# )
    # LDR Conversation
    ldr_conv = ConversationHandler(
        entry_points=[
            CallbackQueryHandler(ldr_request_type_callback, pattern="^(flat_tire|wipers|Drivers_card|other_request)$")
        ],
        states={
            SERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input_ldr)],
            ALLOCATION: [CallbackQueryHandler(allocation_input_ldr)],
            TEAM_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input_ldr)],
            USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input_ldr)],
            DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input_ldr)],
            OTHER_REQUEST_INPUT: [MessageHandler(filters.TEXT & ~filters.COMMAND, ldr_other_request_input)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CallbackQueryHandler(cancel, pattern="cancel")
        ],
        per_user=True,
        conversation_timeout=300  # <--- таймаут 5 минут
    )


    # MFR Conversation
    mfr_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(mfr_callback, pattern="mfr")],
        states={
            # ------------------- Локации -------------------
            ALLOCATION: [
                CallbackQueryHandler(mfr_location_selection, pattern="^(Shyroke|Mykolaiv)$"),
                CallbackQueryHandler(allocation_input_mfr)  # все остальные аллокации
            ],

            # ------------------- Бренды и модели -------------------
            MODEL_SELECTION: [
                CallbackQueryHandler(model_input_mfr, pattern="^(brand_.*|back_to_brands|.*)$")
            ],

            # ------------------- Ввод данных -------------------
            SERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input_mfr)],
            TEAM_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input_mfr)],
            USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input_mfr)],
            DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input_mfr)],
        },
        fallbacks=[
            CommandHandler("cancel", cancel),
            CallbackQueryHandler(cancel, pattern="cancel")
        ],
        per_user=True,
        conversation_timeout=300  # <--- таймаут 5 минут
    )



    
    # Handlers
    app.add_handler(mfr_conv)
    # app.add_handler(other_questions_conv)
    app.add_handler(ldr_conv)
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(start_button_callback, pattern="main_menu"))
    app.add_handler(CallbackQueryHandler(ldr_callback, pattern="ldr"))
    app.add_handler(CallbackQueryHandler(mfr_callback, pattern="mfr"))
    app.add_handler(CallbackQueryHandler(contacts_callback, pattern="contacts"))
    




    app.add_handler(CallbackQueryHandler(cancel, pattern="cancel"))
    app.add_handler(CallbackQueryHandler(contacts_callback, pattern="^contacts$"))
    app.add_handler(CallbackQueryHandler(contact_location_callback, pattern="^contact_shyroke$|^contact_mykolaiv$|^shyroke_carwash$|^shyroke_tire$|^mykolaiv_carwash$|^mykolaiv_tire$|^back$"))


    app.run_polling()

if __name__ == "__main__":
    main()
