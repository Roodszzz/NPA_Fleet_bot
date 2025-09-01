import logging
import os
import re
from io import BytesIO
from datetime import datetime
from unidecode import unidecode
from dotenv import load_dotenv

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, InputFile
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler,
    MessageHandler, ConversationHandler, filters, ContextTypes
)
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from googletrans import Translator

# =================== Загрузка переменных ===================
load_dotenv()
TOKEN = os.getenv("TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

# =================== Логирование ===================
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)

# =================== Состояния ===================
SERIAL, ALLOCATION, TEAM_NUMBER, USER, DESCRIPTION = range(5)
translator = Translator()



# =================== Вспомогательные функции ===================
def get_workbook(report_type="LDR"):
    current_dir = os.path.dirname(__file__)
    if report_type.upper() == "MFR":
        filename = "MFR.xlsx"
    else:
        filename = "LDR.xlsx"
    return load_workbook(os.path.join(current_dir, "excel", filename))

# def get_workbook():
#     current_dir = os.path.dirname(__file__)
#     return load_workbook(os.path.join(current_dir, "excel", "LDR.xlsx"))

def get_logo_bytes():
    current_dir = os.path.dirname(__file__)
    with open(os.path.join(current_dir, "logo", "Drive the NPA way.png"), "rb") as f:
        return BytesIO(f.read())

async def translate_to_en(text: str) -> str:
    result = await translator.translate(text, dest='en')
    return result.text

def set_cell(ws, cell, value):
    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

def auto_adjust(ws, cells):
    for cell in cells:
        value = ws[cell].value
        if value:
            col_letter = ''.join(filter(str.isalpha, cell))
            ws.column_dimensions[col_letter].width = max(
                ws.column_dimensions[col_letter].width or 10,
                len(str(value)) + 2
            )
            ws.row_dimensions[ws[cell].row].height = max(
                ws.row_dimensions[ws[cell].row].height or 15,
                15
            )

# =================== Главное меню ===================
async def main_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("LDR | Lost / Damage report | Втрачено або пошкоджено", callback_data="ldr")],
        [InlineKeyboardButton("MFR | Mechanical failure report | Механічне пошкодження авто", callback_data="mfr")],
        [InlineKeyboardButton("VAR | ВАР", callback_data="var")],
        [InlineKeyboardButton("Contacts | Контакти", callback_data="contacts")],
        [InlineKeyboardButton("Other questions | Інші питання", callback_data="other_questions")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    text = ("🌍 EN\n"
            "Hello! This is the NPA Fleet bot 🚗\n"
            "I can help you create reports for vehicles.\n\n"
            "🌍 UA\n"
            "Привіт! Це бот NPA Fleet 🚗\n"
            "Я допоможу вам створювати звіти по автомобілях.\n\n"
            "What are you interested in today? / Що вас цікавить сьогодні?"
            )
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.message.delete()
        except: pass
        await update.callback_query.message.reply_text(text=text, reply_markup=reply_markup)
    else:
        await update.message.reply_text(text=text, reply_markup=reply_markup)

# =================== Старт ===================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    logo_bytes = get_logo_bytes()
    logo_file = InputFile(logo_bytes, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.message:
        await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)
    elif update.callback_query:
        await update.callback_query.answer()
        await update.callback_query.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

async def start_button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await main_menu(update, context)

# =================== Cancel ===================
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    if update.callback_query:
        await update.callback_query.answer()
        try: await update.callback_query.message.delete()
        except: pass
    await main_menu(update, context)
    return ConversationHandler.END


MANAGERS = {
    "Shyroke": [ADMIN_ID],
    "Mykolaiv": [431019082],
}











# =================== LDR ===================













async def ldr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    keyboard = [
        [InlineKeyboardButton("Flat tire | Пошкоджене колесо", callback_data="flat_tire")],
        [InlineKeyboardButton("Wipers replacement | Заміна дворників", callback_data="wipers")],
        [InlineKeyboardButton("Driver's card | Водійська карта", callback_data="Drivers_card")],
        [InlineKeyboardButton("Other request | Інше звернення", callback_data="other_request")],
        [InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]
    ]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text("Choose request type | Виберіть тип звернення:", reply_markup=InlineKeyboardMarkup(keyboard))

async def ldr_request_type_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "cancel":
        return await cancel(update, context)
    if data == "other_request":
        await query.message.reply_text("You chose: Other request | Ви обрали: Інше звернення.", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]))
        return ConversationHandler.END
    

    context.user_data['wb'] = get_workbook("LDR")
    context.user_data['ws'] = context.user_data['wb'].active

    # context.user_data['wb'] = get_workbook()
    # context.user_data['ws'] = context.user_data['wb'].active
    ws = context.user_data['ws']
    if data == "flat_tire":
        set_cell(ws, "B4", "Flat tyre | Пошкоджене колесо")
    elif data == "wipers":
        set_cell(ws, "B4", "Wipers replacement | Заміна дворників")
    elif data == "Drivers_card":
        set_cell(ws, "B4", "Driver's card | Водійська карта")    
    set_cell(ws, "D4", "Serial / ID / Серійний номер / ID")

    keyboard = [
        [InlineKeyboardButton("Shyroke", callback_data="Shyroke")],
        [InlineKeyboardButton("Mykolaiv", callback_data="Mykolaiv")],
        [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
    ]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text("Select vehicle location | Оберіть локацію автомобіля:", reply_markup=InlineKeyboardMarkup(keyboard))
    return ALLOCATION

# =================== Ввод данных ===================
async def serial_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().upper()  # переводим в верхний регистр
    text = text.replace(" ", "")  # убираем пробелы

    # Если пользователь ввел без дефиса, например AA12, добавим дефис автоматически
    if re.fullmatch(r"[A-Z]{2}\d{2}", text):
        text = text[:2] + "-" + text[2:]

    # проверка формата: две буквы - дефис - две цифры
    if not re.fullmatch(r"[A-Z]{2}-\d{2}", text):
        await update.message.reply_text(
            "❌ Неверный формат номера авто. Формат должен быть: AA-12\nTry again / Спробуйте ще раз:"
        )
        return SERIAL
    

    ws = context.user_data['ws']
    set_cell(ws, "D4", text)
    keyboard = [[InlineKeyboardButton(x, callback_data=x)] for x in ["LOGS","MTT","MDD","TFM","QA","NTS"]]
    keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])
    await update.message.reply_text("Choose Allocation | Оберіть Allocation:", reply_markup=InlineKeyboardMarkup(keyboard))
    return ALLOCATION

async def allocation_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    selection = query.data

    # Проверяем, есть ли workbook
    if 'ws' not in context.user_data:
        await query.message.reply_text(
            "❌ Please start the request from the beginning using /start\n❌ Будь ласка, почніть звернення заново за допомогою /start"
        )
        return ConversationHandler.END

    ws = context.user_data['ws']

    # Обработка отмены
    if selection == "cancel":
        return await cancel(update, context)

    # Выбор локации
    if selection in ["Shyroke", "Mykolaiv"]:
        context.user_data['location'] = selection
        set_cell(ws, "B6", selection)
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter vehicle call sign (e.g. HP-12): \nВведіть внутрішній номер авто (напр. HP-12):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return SERIAL

    # Если NTS, MTT, MDD – спрашиваем номер команды
    if selection.upper() in ["NTS", "MTT", "MDD"]:
        context.user_data['allocation'] = selection.upper()
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            f"Enter team number for {selection.upper()} | Введіть номер команди для {selection.upper()}:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return TEAM_NUMBER

    # Иначе записываем выбор Allocation в ячейку D6
    set_cell(ws, "D6", selection)
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(
        "Enter your full name | Введіть ваше ПІБ:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
    )
    return USER


async def team_number_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ Team number must be a number")
        return TEAM_NUMBER
    ws = context.user_data['ws']
    allocation = context.user_data.get('allocation')
    set_cell(ws, "D6", f"{allocation}-{text}")
    await update.message.reply_text("Enter your full name | Введіть Ім'я та прізвище:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]))
    return USER

async def user_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ You did not enter your name")
        return USER
    user_name_latin = unidecode(text)
    ws = context.user_data['ws']
    set_cell(ws, "F4", user_name_latin)
    set_cell(ws, "A10", user_name_latin)
    location = context.user_data.get('location')
    manager_fa = {"Shyroke":"F.A. Oleksandr Rudnov","Mykolaiv":"F.A. Andriy Padalka"}.get(location,"F.A. Unknown")
    set_cell(ws, "D10", manager_fa)
    set_cell(ws, "B10", datetime.now().strftime("%Y-%m-%d"))
    await update.message.reply_text("Briefly describe the situation | Коротко опишіть ситуацію:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]))
    return DESCRIPTION

# =================== Описание ===================


async def description_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
        return DESCRIPTION

    text_en = await translate_to_en(text)
    ws = context.user_data['ws']
    set_cell(ws, "A9", text_en)
    auto_adjust(ws, ["B4","D4","B6","D6","F4","A10","A9","B10"])

    logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
    img = Image(logo_path)
    img.width, img.height = 396, 72
    ws.add_image(img, "A1")

    plate = ws["D4"].value or "CAR"
    filename = f"LDR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # Отправка только менеджеру по локации
    location = context.user_data.get("location")
    manager_ids = MANAGERS.get(location, [])
    for manager_id in manager_ids:
        file_stream = BytesIO()  # новый поток для каждого менеджера
        ws.parent.save(file_stream)
        file_stream.seek(0)
        await context.bot.send_document(chat_id=manager_id, document=file_stream, filename=filename)
        await context.bot.send_message(chat_id=manager_id, text=f"📄 Новий LDR звіт по локації {location}")

    context.user_data.clear()

    # Пользователю только уведомление
    await update.message.reply_text("✅ Your report has been sent to the manager / Звіт надіслано менеджеру")

    # Отправка приветственного фото с кнопкой
    logo_bytes_start = get_logo_bytes()
    logo_file = InputFile(logo_bytes_start, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

    return ConversationHandler.END





# async def description_input_ldr(update: Update, context: ContextTypes.DEFAULT_TYPE):
#     text = update.message.text.strip()
#     if not text:
#         await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
#         return DESCRIPTION
#     text_en = await translate_to_en(text)
#     ws = context.user_data['ws']
#     set_cell(ws, "A9", text_en)
#     auto_adjust(ws, ["B4","D4","B6","D6","F4","A10","A9","B10"])
#     logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
#     img = Image(logo_path)
#     img.width, img.height = 396, 72
#     ws.add_image(img, "A1")
#     plate = ws["D4"].value or "CAR"
#     filename = f"LDR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
#     file_stream = BytesIO()
#     ws.parent.save(file_stream)
#     file_stream.seek(0)
#     await update.message.reply_document(document=file_stream, filename=filename)
#     await update.message.reply_text("✅ File sent / ✅ Файл відправлено")
#     context.user_data.clear()
#     logo_bytes_start = get_logo_bytes()
#     logo_file = InputFile(logo_bytes_start, filename="logo.png")
#     keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
#     reply_markup = InlineKeyboardMarkup(keyboard)
#     await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)
#     return ConversationHandler.END

# =================== Заглушки ===================
async def generic_stub(update: Update, context: ContextTypes.DEFAULT_TYPE, name="Function"):
    query = update.callback_query
    await query.answer()
    keyboard = [[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]]
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(f"You selected {name}. Function in progress.", reply_markup=InlineKeyboardMarkup(keyboard))


async def var_callback(update: Update, context: ContextTypes.DEFAULT_TYPE): return await generic_stub(update, context, "VAR / ВАР")
async def contacts_callback(update: Update, context: ContextTypes.DEFAULT_TYPE): return await generic_stub(update, context, "Contacts / Контакти")
async def other_questions_callback(update: Update, context: ContextTypes.DEFAULT_TYPE): return await generic_stub(update, context, "Other questions / Інші питання")



#=====================================================MFR==================================================

















ALLOCATION, MODEL_SELECTION, SERIAL, TEAM_NUMBER, USER, DESCRIPTION = range(6)


# Начало MFR запроса — спрашиваем локацию
async def mfr_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    # Создаем workbook для MFR
    context.user_data['wb'] = get_workbook("MFR")
    context.user_data['ws'] = context.user_data['wb'].active
    ws = context.user_data['ws']

    # Тип запроса в Excel
    set_cell(ws, "F6", "Serial / ID / Серійний номер / ID")

    # ------------------- Кнопки локации -------------------
    keyboard = [
        [InlineKeyboardButton("Shyroke", callback_data="Shyroke")],
        [InlineKeyboardButton("Mykolaiv", callback_data="Mykolaiv")],
        [InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]
    ]
    # -------------------------------------------------------

    try: 
        await query.message.delete()
    except: 
        pass

    await query.message.reply_text(
        "Select vehicle location | Оберіть локацію автомобіля:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ALLOCATION




# Выбор локации
async def mfr_location_selection(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    location = query.data

    if location == "cancel":
        return await cancel(update, context)

    ws = context.user_data['ws']
    set_cell(ws, "C9", location)
    context.user_data['location'] = location

    try: 
        await query.message.delete()
    except: 
        pass

    # ------------------- Кнопки с моделями авто -------------------
    keyboard = [
        [InlineKeyboardButton("Toyota Hilux", callback_data="Toyota Hilux")],
        [InlineKeyboardButton("Toyota Land Cruiser", callback_data="Toyota Land Cruiser")],
        [InlineKeyboardButton("Ford Transit", callback_data="Ford Transit")],
        [InlineKeyboardButton("Ford Ranger", callback_data="Ford Ranger")],
        [InlineKeyboardButton("Mitsubishi L200", callback_data="Mitsubishi L200")],
        [InlineKeyboardButton("Volkswagen", callback_data="Volkswagen")],
        [InlineKeyboardButton("Renault Duster", callback_data="Renault Duster")],
        [InlineKeyboardButton("❌ Cancel / Відмінити", callback_data="cancel")]
    ]
    # ---------------------------------------------------------------

    await query.message.reply_text(
        "Select car model | Оберіть модель авто:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return MODEL_SELECTION




# Выбор модели авто
async def model_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    model_name = query.data

    if model_name == "cancel":
        return await cancel(update, context)

    ws = context.user_data['ws']
    set_cell(ws, "C6", model_name)  # записываем выбранную модель в C6

    try: 
        await query.message.delete()
    except: 
        pass

    await query.message.reply_text(
        "Enter vehicle call sign (e.g. HP-12): \nВведіть внутрішній номер авто (напр. HP-12):",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
    )
    return SERIAL



async def serial_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().upper()
    text = text.replace(" ", "")

    if re.fullmatch(r"[A-Z]{2}\d{2}", text):
        text = text[:2] + "-" + text[2:]

    if not re.fullmatch(r"[A-Z]{2}-\d{2}", text):
        await update.message.reply_text(
            "❌ Неверный формат номера авто. Формат должен быть: AA-12\nTry again | Спробуйте ще раз:"
        )
        return SERIAL

    ws = context.user_data['ws']
    set_cell(ws, "F6", text)

    keyboard = [[InlineKeyboardButton(x, callback_data=x)] for x in ["LOGS","MTT","MDD","TFM","QA","NTS"]]
    keyboard.append([InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")])
    await update.message.reply_text("Choose Allocation | Оберіть Allocation:", reply_markup=InlineKeyboardMarkup(keyboard))
    return ALLOCATION


async def allocation_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if 'ws' not in context.user_data:
        await query.message.reply_text("❌ Please start from /start")
        return ConversationHandler.END

    ws = context.user_data['ws']
    selection = query.data

    if selection == "cancel":
        return await cancel(update, context)

    if selection in ["Shyroke", "Mykolaiv"]:
        context.user_data['location'] = selection
        set_cell(ws, "C9", selection)
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            "Enter vehicle call sign (e.g. HP-12) | Введіть внутрішній номер авто (напр. HP-12):",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return SERIAL

    if selection.upper() in ["NTS", "MTT", "MDD"]:
        context.user_data['allocation'] = selection.upper()
        try: await query.message.delete()
        except: pass
        await query.message.reply_text(
            f"Enter team number for {selection.upper()} | Введіть номер команди для {selection.upper()}:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
        )
        return TEAM_NUMBER

    set_cell(ws, "F9", selection)
    try: await query.message.delete()
    except: pass
    await query.message.reply_text(
        "Enter your full name | Введіть ваше ім'я та прізвище:",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ Cancel | Відмінити", callback_data="cancel")]])
    )
    return USER


async def team_number_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit():
        await update.message.reply_text("❌ Team number must be a number")
        return TEAM_NUMBER
    ws = context.user_data['ws']
    allocation = context.user_data.get('allocation')
    set_cell(ws, "F9", f"{allocation}-{text}")
    await update.message.reply_text("Enter your full name | Введіть ПІБ:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]))
    return USER

async def user_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ You did not enter your name")
        return USER
    user_name_latin = unidecode(text)
    ws = context.user_data['ws']
    set_cell(ws, "I6", user_name_latin)
    set_cell(ws, "B23", user_name_latin)
    location = context.user_data.get('location')
    manager_fa = {"Shyroke":"F.A. Oleksandr Rudnov","Mykolaiv":"F.A. Andriy Padalka"}.get(location,"F.A. Unknown")
    set_cell(ws, "F23", manager_fa)
    set_cell(ws, "C23", datetime.now().strftime("%Y-%m-%d"))
    set_cell(ws, "F12", datetime.now().strftime("%Y-%m-%d"))
    await update.message.reply_text("Briefly describe the situation | Коротко опишіть ситуацію:", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("Cancel / Відмінити", callback_data="cancel")]]))
    return DESCRIPTION

# =================== Описание ===================
# async def description_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
#     text = update.message.text.strip()
#     if not text:
#         await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
#         return DESCRIPTION
#     text_en = await translate_to_en(text)
#     ws = context.user_data['ws']
#     set_cell(ws, "B15", text_en)
#     auto_adjust(ws, ["F5", "C5", "C8", "F8", "I5", "F22", "C22", "B15"])
#     logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
#     img = Image(logo_path)
#     img.width, img.height = 396, 72
#     ws.add_image(img, "A1")
#     plate = ws["F5"].value or "CAR"
#     filename = f"MFR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
#     file_stream = BytesIO()
#     ws.parent.save(file_stream)
#     file_stream.seek(0)
#     await update.message.reply_document(document=file_stream, filename=filename)
#     await update.message.reply_text("✅ File sent / ✅ Файл відправлено")
#     context.user_data.clear()
#     logo_bytes_start = get_logo_bytes()
#     logo_file = InputFile(logo_bytes_start, filename="logo.png")
#     keyboard = [[InlineKeyboardButton("Start / Почати", callback_data="main_menu")]]
#     reply_markup = InlineKeyboardMarkup(keyboard)
#     await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)
#     return ConversationHandler.END



# async def description_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
#     text = update.message.text.strip()
#     if not text:
#         await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
#         return DESCRIPTION

#     text_en = await translate_to_en(text)
#     ws = context.user_data['ws']
#     set_cell(ws, "B15", text_en)
#     auto_adjust(ws, ["F5", "C5", "C8", "F8", "I5", "F22", "C22", "B15"])

#     logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
#     img = Image(logo_path)
#     img.width, img.height = 396, 72
#     ws.add_image(img, "A1")

#     plate = ws["F5"].value or "CAR"
#     filename = f"MFR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

#     # Отправка только менеджеру по локации
#     location = context.user_data.get("location")
#     manager_ids = MANAGERS.get(location, [])
#     for manager_id in manager_ids:
#         file_stream = BytesIO()  # новый поток для каждого менеджера
#         ws.parent.save(file_stream)
#         file_stream.seek(0)
#         await context.bot.send_document(chat_id=manager_id, document=file_stream, filename=filename)
#         await context.bot.send_message(chat_id=manager_id, text=f"📄 Новий звіт по локації {location}")

#     context.user_data.clear()

#     # Пользователю только уведомление
#     await update.message.reply_text("✅ Your report has been sent to the manager / Звіт надіслано менеджеру")
#     return ConversationHandler.END

# async def description_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
#     text = update.message.text.strip()
#     if not text:
#         await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
#         return DESCRIPTION

#     text_en = await translate_to_en(text)
#     ws = context.user_data['ws']
#     set_cell(ws, "B15", text_en)
#     auto_adjust(ws, ["F5", "C5", "C8", "F8", "I5", "F22", "C22", "B15"])

#     logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
#     img = Image(logo_path)
#     img.width, img.height = 396, 72
#     ws.add_image(img, "B1")

#     plate = ws["F5"].value or "CAR"
#     filename = f"MFR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

#     # Отправка только менеджеру по локации
#     location = context.user_data.get("location")
#     manager_ids = MANAGERS.get(location, [])
#     for manager_id in manager_ids:
#         file_stream = BytesIO()  # создаем новый поток для каждого менеджера
#         ws.parent.save(file_stream)
#         file_stream.seek(0)
#         await context.bot.send_document(chat_id=manager_id, document=file_stream, filename=filename)
#         await context.bot.send_message(chat_id=manager_id, text=f"📄 Новий MFR звіт по локації {location}")

#     context.user_data.clear()

#     # Пользователю только уведомление
#     await update.message.reply_text("✅ Your report has been sent to the manager / Звіт надіслано менеджеру")

#     # Отправка приветственного фото с кнопкой
#     logo_bytes_start = get_logo_bytes()
#     logo_file = InputFile(logo_bytes_start, filename="logo.png")
#     keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
#     reply_markup = InlineKeyboardMarkup(keyboard)
#     await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

#     return ConversationHandler.END


# async def description_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
#     text = update.message.text.strip()
#     if not text:
#         await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
#         return DESCRIPTION

#     text_en = await translate_to_en(text)
#     ws = context.user_data['ws']

#     # --- Текст с переносом ---
#     cell = ws["B15"]
#     cell.value = text_en
#     cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

#     # --- Фиксируем ширину колонки ---
#     ws.column_dimensions["B"].width = 50  # можно подкорректировать ширину

#     # --- Подгоняем высоту строки ---
#     approx_char_per_line = 50  # примерно символов в строке
#     lines = (len(text_en) // approx_char_per_line) + 1
#     ws.row_dimensions[cell.row].height = lines * 15  # высота строки

#     # --- Остальные ячейки авто ---
#     auto_adjust(ws, ["F5", "C5", "C8", "F8", "I5", "F22", "C22"])

#     # --- Лого ---
#     logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
#     img = Image(logo_path)
#     img.width, img.height = 396, 72
#     ws.add_image(img, "B1")

#     plate = ws["F5"].value or "CAR"
#     filename = f"MFR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

#     # --- Отправка менеджеру ---
#     location = context.user_data.get("location")
#     manager_ids = MANAGERS.get(location, [])
#     for manager_id in manager_ids:
#         file_stream = BytesIO()
#         ws.parent.save(file_stream)
#         file_stream.seek(0)
#         await context.bot.send_document(chat_id=manager_id, document=file_stream, filename=filename)
#         await context.bot.send_message(chat_id=manager_id, text=f"📄 Новий MFR звіт по локації {location}")

#     context.user_data.clear()
#     await update.message.reply_text("✅ Your report has been sent to the manager / Звіт надіслано менеджеру")

#     # --- Приветственное фото ---
#     logo_bytes_start = get_logo_bytes()
#     logo_file = InputFile(logo_bytes_start, filename="logo.png")
#     keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
#     reply_markup = InlineKeyboardMarkup(keyboard)
#     await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

#     return ConversationHandler.END


async def description_input_mfr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("❌ Describe the situation / ❌ Опишіть ситуацію")
        return DESCRIPTION

    text_en = await translate_to_en(text)
    ws = context.user_data['ws']

    # --- Определяем диапазон для текстового блока ---
    start_row = 16
    end_row = 21
    col = "B"
    max_chars_per_line = 150  # примерная длина строки

    # --- Разбиваем текст на строки ---
    words = text_en.split()
    lines = []
    current_line = ""
    for word in words:
        if len(current_line) + len(word) + 1 <= max_chars_per_line:
            current_line += (" " if current_line else "") + word
        else:
            lines.append(current_line)
            current_line = word
    if current_line:
        lines.append(current_line)

    # --- Записываем в ячейки ---
    for i, row in enumerate(range(start_row, end_row + 1)):
        if i < len(lines):
            ws[f"{col}{row}"].value = lines[i]
            ws[f"{col}{row}"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        else:
            ws[f"{col}{row}"].value = ""
    
    # --- Остальные ячейки авто ---
    auto_adjust(ws, ["F6", "C6", "C9", "F9", "I6", "F23", "C23"])

    # --- Лого ---
    logo_path = os.path.join(os.path.dirname(__file__), "logo", "Лого ексель.png")
    img = Image(logo_path)
    img.width, img.height = 396, 72
    ws.add_image(img, "B2")

    plate = ws["F6"].value or "CAR"
    filename = f"MFR_{plate}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    # --- Отправка менеджеру ---
    location = context.user_data.get("location")
    manager_ids = MANAGERS.get(location, [])
    for manager_id in manager_ids:
        file_stream = BytesIO()
        ws.parent.save(file_stream)
        file_stream.seek(0)
        await context.bot.send_document(chat_id=manager_id, document=file_stream, filename=filename)
        await context.bot.send_message(chat_id=manager_id, text=f"📄 Новий MFR звіт по локації {location}")

    context.user_data.clear()
    await update.message.reply_text("✅ Your report has been sent to the manager / Звіт надіслано менеджеру")

    # --- Приветственное фото ---
    logo_bytes_start = get_logo_bytes()
    logo_file = InputFile(logo_bytes_start, filename="logo.png")
    keyboard = [[InlineKeyboardButton("Start | Почати", callback_data="main_menu")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_photo(photo=logo_file, caption="Welcome to NPA Fleet bot 🚗", reply_markup=reply_markup)

    return ConversationHandler.END



























# =================== Main ===================
def main():
    app = Application.builder().token(TOKEN).build()

# LDR Conversation


    ldr_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(ldr_request_type_callback, pattern="^(flat_tire|wipers|Drivers_card|other_request)$")],
        states={
            SERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input_ldr)],
            ALLOCATION: [CallbackQueryHandler(allocation_input_ldr)],
            TEAM_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input_ldr)],
            USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input_ldr)],
            DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input_ldr)],
        },
        fallbacks=[CommandHandler("cancel", cancel), CallbackQueryHandler(cancel, pattern="cancel")],
        per_user=True
    )
# MFR Conversation


    mfr_conv = ConversationHandler(
        entry_points=[CallbackQueryHandler(mfr_callback, pattern="mfr")],
        states={
            ALLOCATION: [
                CallbackQueryHandler(mfr_location_selection, pattern="^(Shyroke|Mykolaiv)$"),
                CallbackQueryHandler(allocation_input_mfr, pattern="^(LOGS|MTT|MDD|TFM|QA|NTS)$")
            ],
            MODEL_SELECTION: [
                CallbackQueryHandler(model_input_mfr, pattern="^(Toyota Hilux|Toyota Land Cruiser|Ford Transit|Ford Ranger|Mitsubishi L200|Volkswagen|Renault Duster)$")
            ],
            SERIAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, serial_input_mfr)],
            TEAM_NUMBER: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_number_input_mfr)],
            USER: [MessageHandler(filters.TEXT & ~filters.COMMAND, user_input_mfr)],
            DESCRIPTION: [MessageHandler(filters.TEXT & ~filters.COMMAND, description_input_mfr)],
        },
        fallbacks=[CommandHandler("cancel", cancel), CallbackQueryHandler(cancel, pattern="cancel")],
        per_user=True
    )







    
    # Handlers
    app.add_handler(mfr_conv)

    app.add_handler(ldr_conv)
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CallbackQueryHandler(start_button_callback, pattern="main_menu"))
    app.add_handler(CallbackQueryHandler(ldr_callback, pattern="ldr"))
    app.add_handler(CallbackQueryHandler(mfr_callback, pattern="mfr"))
    app.add_handler(CallbackQueryHandler(var_callback, pattern="var"))
    app.add_handler(CallbackQueryHandler(contacts_callback, pattern="contacts"))
    app.add_handler(CallbackQueryHandler(other_questions_callback, pattern="other_questions"))
    app.add_handler(CallbackQueryHandler(cancel, pattern="cancel"))

    app.run_polling()

if __name__ == "__main__":
    main()
